import json
import time
import io
import openpyxl
from telegram import KeyboardButton
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import CallbackQueryHandler
import asyncio
import os
import psycopg2
from telegram import InputMediaPhoto

DATABASE_URL = os.getenv("DATABASE_URL")

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()   # 🔥 SHU MUHIM
cur.execute("""
CREATE TABLE IF NOT EXISTS products (
    id SERIAL PRIMARY KEY,
    photo TEXT,
    gender TEXT,
    origin TEXT,
    season TEXT,
    category TEXT,
    name TEXT,
    size TEXT,
    price TEXT,
    count INTEGER,
    reserved INTEGER DEFAULT 0
)
""")

cur.execute("UPDATE products SET reserved = 0")
conn.commit()

cur.execute("DELETE FROM products WHERE photo IS NULL OR photo = ''")
conn.commit()

cur.execute("""
CREATE TABLE IF NOT EXISTS photos (
    id SERIAL PRIMARY KEY,
    file_id TEXT NOT NULL,
    created_at FLOAT
)
""")
conn.commit()

cur.execute("""
ALTER TABLE products ADD COLUMN IF NOT EXISTS cost INTEGER DEFAULT 0
""")
conn.commit()


cur.execute("""
CREATE TABLE IF NOT EXISTS orders (
    id SERIAL PRIMARY KEY,
    user_id BIGINT,
    cart TEXT,
    location TEXT,
    phone TEXT,
    total INTEGER,
    status TEXT,
    time FLOAT
)
""")

conn.commit()



cur.execute("""
CREATE TABLE IF NOT EXISTS users (
    user_id BIGINT PRIMARY KEY
)
""")
conn.commit()

cur.execute("""
ALTER TABLE users ADD COLUMN IF NOT EXISTS created_at FLOAT
""")
conn.commit()

TOKEN = os.getenv("TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

product_locks = {}
products = []
carts = {}
order_id_counter = 1
ORIGINS = [
    "🇺🇿 Vodiy",
    "🇨🇳 Xitoy",
    "🇹🇷 Turkiya",
    "🏭 8-mart fabrika"
]
def get_filter_menu(user_data):
    g = user_data.get("filter_gender", "-")
    o = user_data.get("filter_origin", "-")
    s = user_data.get("filter_season", "-")

    return InlineKeyboardMarkup([

        # 🔥 JINS LABEL
        [InlineKeyboardButton(f"Jins: {g}", callback_data="empty")],

        [
            InlineKeyboardButton(f"👦 O‘g‘il {'✅' if g=='o‘g‘il' else ''}", callback_data="g_o‘g‘il"),
            InlineKeyboardButton(f"👧 Qiz {'✅' if g=='qiz' else ''}", callback_data="g_qiz"),
        ],

        # 🔥 FABRIKA LABEL
        [InlineKeyboardButton(f"Fabrika: {o}", callback_data="empty")],

        [
            InlineKeyboardButton(f"🇺🇿 Vodiy {'✅' if o=='Vodiy' else ''}", callback_data="o_Vodiy"),
            InlineKeyboardButton(f"🇨🇳 Xitoy {'✅' if o=='Xitoy' else ''}", callback_data="o_Xitoy"),
        ],
        [
            InlineKeyboardButton(f"🇹🇷 Turkiya {'✅' if o=='Turkiya' else ''}", callback_data="o_Turkiya"),
            InlineKeyboardButton(f"🏭 8-mart {'✅' if o=='8-mart fabrika' else ''}", callback_data="o_8-mart fabrika"),
        ],

        # 🔥 FASL LABEL
        [InlineKeyboardButton(f"Fasl: {s}", callback_data="empty")],

        [
            InlineKeyboardButton(f"☀️ Yozgi {'✅' if s=='Yozgi' else ''}", callback_data="s_Yozgi"),
            InlineKeyboardButton(f"❄️ Qishki {'✅' if s=='Qishki' else ''}", callback_data="s_Qishki"),
        ],
        [
            InlineKeyboardButton(f"🌸 Bahor {'✅' if s=='Bahor' else ''}", callback_data="s_Bahor"),
            InlineKeyboardButton(f"🍂 Kuz {'✅' if s=='Kuz' else ''}", callback_data="s_Kuz"),
        ],

        [
            InlineKeyboardButton("✅ Tanlash", callback_data="apply"),
            InlineKeyboardButton("🔄 Tozalash", callback_data="reset"),
        ]
    ])
def get_category_buttons(context):
    return [
        [f"👕 2 talik ({count_products(context,lambda p: p['category'].lower()=='2 talik kiyim')})",
         f"👕 3 talik ({count_products(context,lambda p: p['category'].lower()=='3 talik kiyim')})",
         f"👕 Futbolka ({count_products(context,lambda p: p['category'].lower()=='futbolka')})"],

         [f"👖 Shim ({count_products(context,lambda p: p['category'].lower()=='shim')})",
          f"🧥 Qalin ({count_products(context,lambda p: p['category'].lower()=='qalin kiyim')})",
         f"🩳 Shortik ({count_products(context,lambda p: p['category'].lower()=='shortik')})"],

        [f"👟 Oyoq ({count_products(context,lambda p: p['category'].lower()=='oyoq kiyim')})",
         f"🧢 Bosh ({count_products(context,lambda p: p['category'].lower()=='bosh kiyim')})",
         f"🩲 Ichki ({count_products(context,lambda p: p['category'].lower()=='ichki kiyim')})"],

        ["🔙 Orqaga", "🏠 Bosh menyu"]
    ]

def filter_check(p, context):

    # 🔥 gender
    g = context.user_data.get("filter_gender")
    if g:
        if g.lower() not in str(p.get("gender", "")).lower():
            return False

    # 🔥 origin
    o = context.user_data.get("filter_origin")
    if o:
        if o.lower() not in str(p.get("origin", "")).lower():
            return False

    # 🔥 category (YENGIL VA ISHONCHLI)
    c = context.user_data.get("filter_category")
    if c:
        user_cat = str(c).strip().lower()
        prod_cat = str(p.get("category", "")).strip().lower()

        if not prod_cat:
            return False

        # 🔥 faqat o‘xshashlikni tekshiramiz (soft)
        if user_cat not in prod_cat:
            return False

    # 🔥 season (TOZA)
    s = context.user_data.get("filter_season")
    if s:
        season = str(s).strip().lower()

        p_seasons = p.get("season", [])

        # har doim listga aylantiramiz
        if not isinstance(p_seasons, list):
            p_seasons = str(p_seasons).split(",")

        p_seasons = [str(x).strip().lower() for x in p_seasons]

        if season not in p_seasons:
            return False

    # 🔥 size
    if context.user_data.get("filter_size"):
        size_text = context.user_data.get("filter_size")
        if str(size_text).isdigit():
            size = int(size_text)
            raw = str(p.get("size") or "").lower().replace("sm", "").strip()

            if raw == "":
                print(f"DEBUG: {p['name']} - o'lcham bo'sh")
                return False

            if "-" in raw:
                parts = raw.split("-")
                if len(parts) >= 2 and parts[0].strip().isdigit() and parts[1].strip().isdigit():
                    s1, s2 = int(parts[0]), int(parts[1])
                    if not (s1 <= size <= s2):
                        print(f"DEBUG: {p['name']} - o'lcham diapazonga tushmadi ({s1}-{s2})")
                        return False
                else: return False
            else:
                if not raw.isdigit(): return False
                p_size = int(raw)
                if abs(p_size - size) > 1:
                    print(f"DEBUG: {p['name']} - o'lcham mos kelmadi (P:{p_size}, U:{size})")
                    return False

    # 🔥 mavjudlik (Eng ko'p xato shu yerda bo'ladi)
    available = p.get("count", 0) - p.get("reserved", 0)

    if available <= 0:
        return False

    # Agar barcha shartlardan o'tsa
    return True 
def clean_cart(user_id, context=None):
    
    now = time.time()

    # 🔥 BUYURTMA BOSILGAN BO‘LSA — TO‘XTAT
    if context and context.user_data.get("order_started"):
        return carts.get(user_id, {})

    cart = carts.get(user_id, {})
    new_cart = {}

    for pid, item in cart.items():
        if now - item["time"] < 7200:
            new_cart[pid] = item
        else:
            p = next((x for x in products if x["id"] == int(pid)), None)
            if p:
                p["reserved"] = max(0, p.get("reserved", 0) - item["qty"])

    carts[user_id] = new_cart
    return new_cart

def load_products_from_db():
    global products

    cur.execute("SELECT * FROM products")
    rows = cur.fetchall()

    products = []

    for r in rows:
        products.append({
            "id": r[0],
            "photo": r[1],
            "gender": r[2],
            "origin": r[3],
            "season": r[4].split(",") if r[4] else [],
            "category": r[5],
            "name": r[6],
            "size": r[7],
            "price": r[8],
            "count": r[9],
            "reserved": r[10]
        })
def count_products(context, filter_func=None):
    total = 0
    for p in products:
        available = p["count"] - p.get("reserved", 0)

        if available > 0 and filter_check(p, context):
            if not filter_func or filter_func(p):
                total += available

    return total


ADMIN_MENU = ReplyKeyboardMarkup(
    [
        ["📊 Statistika", "📦 Buyurtmalar"],
        [ "📢 Reklama"],
        ["🏠 Bosh menyu"]
    ],
    resize_keyboard=True
)

MAIN_MENU = ReplyKeyboardMarkup(
    [
        ["🔍 Qidirish"],
        ["🛍 Kiyimlar", "🧺 Savat"],
        ["ℹ️ Yordam"]
    ],
    resize_keyboard=True
)

BACK_BUTTON = ["🔙 Orqaga"]
HOME_BUTTON = ["🏠 Bosh menyu"]
CART_BUTTON = ["🧺 Savat"]

CATEGORIES = [
    "👕 2 talik kiyim (dvoyka)",
    "👕 3 talik kiyim (troyka)",
    "👕 Futbolka(qizlarga ko‘ylak ham)",
    "👖 Shim",
    "🧥 Qalin kiyim",
    "🩳 Shortik(qizlarga yubka)",
    "👟 oyoq kiyim",
    "🧢 Bosh kiyim",
    "🩲 Ichki kiyim"
]

async def show_products(update, context, products, ADMIN_ID):

    page = context.user_data.get("page", 0)
    start = page * 4
    end = start + 4

    chunk = products[start:end]

    if not chunk:
        await update.message.reply_text("❌ Hech narsa topilmadi")
        return

    # 🔥 MEDIA VA VALID MAHSULOTLAR YIG'ISH
    media = []
    valid_products = []

    for i, p in enumerate(chunk):
        if not p.get("photo"):
            continue  # rasm yo'q bo'lsa skip

        media.append(
            InputMediaPhoto(
                media=p.get("photo"),
                caption=f"{i+1}) {p.get('size')}"
            )
        )
        valid_products.append(p)

    # 🔥 ALBUM YUBORISH
    if media:
        await update.message.reply_media_group(media)
    else:
        await update.message.reply_text("❌ Rasm topilmadi")
        return

    # 🔽 INFO + TUGMALAR
    for i, p in enumerate(valid_products):
        keyboard = [
            [InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p.get('id')}")]
        ]

        if update.effective_user.id == ADMIN_ID:
            keyboard.append([
                InlineKeyboardButton("✏️ Edit", callback_data=f"edit_{p.get('id')}"),
                InlineKeyboardButton("🗑 O‘chirish", callback_data=f"delete_{p.get('id')}")
            ])

        await update.message.reply_text(
            f"{i+1}) {p.get('name')}\n📏 {p.get('size')} sm\n💰 {p.get('price')}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # 🔽 NEXT / PREV
    nav = []

    if start > 0:
        nav.append(InlineKeyboardButton("⬅️ Orqaga", callback_data="prev"))

    if end < len(products):
        nav.append(InlineKeyboardButton("➡️ Keyingi", callback_data="next"))

    if nav:
        await update.message.reply_text(
            "📄 Sahifa",
            reply_markup=InlineKeyboardMarkup([nav])
        )

async def start_photo_flow(context, chat_id, file_id):
    # 🔥 bu photo_handler ichidagi boshlanish logikasi
    context.user_data.clear()
    context.user_data["photo"] = file_id
    context.user_data["step"] = "gender"

    keyboard = [["👦 O‘g‘il", "👧 Qiz"]]
    await context.bot.send_message(
        chat_id=chat_id,
        text="Kim uchun?",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
# START
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    context.user_data.clear()

    user_id = update.effective_user.id

    cur.execute("""
    INSERT INTO users (user_id, created_at)
    VALUES (%s, %s)
    ON CONFLICT (user_id) DO NOTHING
    """, (user_id, time.time()))
    conn.commit()

    load_products_from_db()

    # 🔥 BIR MARTA JAVOB BERAMIZ
    if user_id == ADMIN_ID:
        await update.message.reply_text(
            "👑 Admin panel, Assalomu aleykum AZIZJON AHMADOVICH " ,
            reply_markup=ADMIN_MENU
        )
    else:
        await update.message.reply_text(
            "Assalomu alaykum 👋",
            reply_markup=MAIN_MENU
        )
# RASM QABUL (ADMIN)
async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if user_id != ADMIN_ID and not update.message.from_user.is_bot:
        return

    # /get_id rejimi — rasm saqlanib raqam qaytaradi
    if context.user_data.get("get_id_mode"):
        file_id = update.message.photo[-1].file_id
        cur.execute(
            "INSERT INTO photos (file_id, created_at) VALUES (%s, %s) RETURNING id",
            (file_id, time.time())
        )
        photo_num = cur.fetchone()[0]
        conn.commit()
        await update.message.reply_text(
            f"✅ Rasm saqlandi!\n\n"
            f"📋 Raqam: <b>#{photo_num}</b>\n\n"
            f"Excel jadvalga <code>{photo_num}</code> raqamini yozing",
            parse_mode="HTML"
        )
        return

    # Oddiy mahsulot qo'shish rejimi
    context.user_data.clear()
    context.user_data["photo"] = update.message.photo[-1].file_id
    context.user_data["step"] = "gender"

    keyboard = [["👦 O'g'il", "👧 Qiz"]]
    await update.message.reply_text(
        "Kim uchun?",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )


# /get_id buyrug'i
async def get_id_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    context.user_data["get_id_mode"] = True
    await update.message.reply_text(
        "📸 Endi rasmlarni ketma-ket yuboring\n"
        "Har bir rasm uchun raqam qaytaraman (#1, #2, #3...)\n"
        "Shu raqamni Excel ga yozing.\n\n"
        "Tugagach /done yuboring.",
        reply_markup=ReplyKeyboardMarkup([["🏠 Bosh menyu"]], resize_keyboard=True)
    )


# /done buyrug'i
async def done_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("get_id_mode", None)
    await update.message.reply_text(
        "✅ file_id rejimi tugadi.",
        reply_markup=ADMIN_MENU if update.effective_user.id == ADMIN_ID else MAIN_MENU
    )


# Excel import validatsiya konstantlari
FASL_OPTIONS = {"yozgi", "qishki", "bahor", "kuz"}
VALID_CATEGORIES = {
    "2 talik kiyim", "3 talik kiyim", "futbolka", "shim",
    "qalin kiyim", "shortik", "oyoq kiyim", "bosh kiyim", "ichki kiyim"
}


async def excel_import_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    doc = update.message.document
    if not doc or not doc.file_name.endswith(".xlsx"):
        return

    await update.message.reply_text("⏳ Excel o'qilmoqda...")

    file = await context.bot.get_file(doc.file_id)
    file_bytes = await file.download_as_bytearray()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb["Mahsulotlar"]
    except Exception as e:
        await update.message.reply_text(f"❌ Fayl o'qilmadi: {e}")
        return

    added = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row[0] and not row[1]:
            continue
        if row[0] and "NAMUNA" in str(row[0]).upper():
            continue

        photo_raw = str(row[0]).strip() if row[0] else ""
        name      = str(row[1]).strip() if row[1] else ""

        # Raqam bo'lsa — photos jadvaldan file_id olish
        if photo_raw.isdigit():
            cur.execute("SELECT file_id FROM photos WHERE id = %s", (int(photo_raw),))
            ph_row = cur.fetchone()
            if not ph_row:
                errors.append(f"Qator {row_num}: #{photo_raw} raqamli rasm topilmadi")
                continue
            photo = ph_row[0]
        else:
            photo = photo_raw  # to'g'ridan file_id yozilgan bo'lsa ham ishlaydi
        gender    = str(row[2]).strip() if row[2] else ""
        origin    = str(row[3]).strip() if row[3] else ""
        season    = str(row[4]).strip() if row[4] else ""
        category  = str(row[5]).strip().lower() if row[5] else ""
        size      = str(row[6]).strip() if row[6] else ""
        price_raw = str(row[7]).strip() if row[7] else ""
        cost_raw  = str(row[8]).strip() if row[8] else "0"
        count_raw = str(row[9]).strip() if row[9] else ""

        missing = []
        if not photo_raw: missing.append("photo raqami")
        if not name:      missing.append("nom")
        if not gender:    missing.append("jins")
        if not origin:    missing.append("fabrika")
        if not season:    missing.append("fasl")
        if not category:  missing.append("kategoriya")
        if not size:      missing.append("razmer")
        if not price_raw: missing.append("narx")
        if not count_raw: missing.append("soni")

        if missing:
            errors.append(f"Qator {row_num}: {', '.join(missing)} yetishmaydi")
            continue

        try:
            price_int = int("".join(filter(str.isdigit, price_raw)))
            price_str = f"{price_int:,}".replace(",", " ") + " so'm"
        except:
            errors.append(f"Qator {row_num}: narx noto'g'ri ({price_raw})")
            continue

        try:
            cost_int = int("".join(filter(str.isdigit, cost_raw))) if cost_raw and cost_raw != "0" else 0
        except:
            cost_int = 0

        try:
            count_int = int("".join(filter(str.isdigit, count_raw)))
        except:
            errors.append(f"Qator {row_num}: soni noto'g'ri ({count_raw})")
            continue

        season_parts = [s.strip().capitalize() for s in season.replace(",", " ").split()]
        valid_seasons = [s for s in season_parts if s.lower() in FASL_OPTIONS]
        if not valid_seasons:
            errors.append(f"Qator {row_num}: fasl noto'g'ri ({season})")
            continue
        season_db = ",".join(valid_seasons)

        if category not in VALID_CATEGORIES:
            errors.append(f"Qator {row_num}: kategoriya noto'g'ri ({category})")
            continue

        try:
            cur.execute("""
                INSERT INTO products
                    (photo, gender, origin, season, category, name, size, price, count, reserved, cost)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s)
            """, (
                photo, gender, origin, season_db,
                category, name, size, price_str,
                count_int, cost_int
            ))
            added += 1
        except Exception as e:
            errors.append(f"Qator {row_num}: DB xato — {e}")
            conn.rollback()
            continue

    conn.commit()
    load_products_from_db()

    msg = f"✅ {added} ta mahsulot qo'shildi!"
    if errors:
        err_text = "\n".join(errors[:10])
        if len(errors) > 10:
            err_text += f"\n... va yana {len(errors)-10} ta xato"
        msg += f"\n\n⚠️ Xatolar:\n{err_text}"

    await update.message.reply_text(msg)


async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not update.message:
            return

        text = update.message.text
        # ===== ADMIN FLOW =====
        if context.user_data.get("step") == "gender":
            gender = text.replace("👦 ", "").replace("👧 ", "")
            context.user_data["gender"] = gender

            # 🔥 YANGI STEP
            context.user_data["step"] = "origin"

            keyboard = [
                ["🇺🇿 Vodiy", "🇨🇳 Xitoy"],
                ["🇹🇷 Turkiya", "🏭 8-mart fabrika"]
            ]

            await update.message.reply_text(
                "Qaysi ishlab chiqarish:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        


        elif context.user_data.get("step") == "broadcast":
            msg = text

            cur.execute("SELECT user_id FROM users")
            users = cur.fetchall()

            count = 0

            for u in users:
                try:
                    await context.bot.send_message(chat_id=u[0], text=msg)
                    count += 1
                except:
                    pass

            await update.message.reply_text(f"✅ {count} ta userga yuborildi")

            context.user_data.clear()

        elif text == "🔍 Qidirish":
            context.user_data.clear()

            await update.message.reply_text(
                "🔎 Tanlang:\n\nJins: -\nFabrika: -\nFasl: -\n\n",
                reply_markup=get_filter_menu(context.user_data)
            )

            context.user_data["step"] = "inline_all"

        elif text == "📢 Reklama":
            if update.effective_user.id != ADMIN_ID:
                return

            context.user_data["step"] = "broadcast"
            await update.message.reply_text("📢 Yuboriladigan matnni yozing:")

        elif text == "📦 Buyurtmalar":
            if update.effective_user.id != ADMIN_ID:
                return

            cur.execute("SELECT id, total, status FROM orders ORDER BY id DESC LIMIT 10")
            rows = cur.fetchall()

            if not rows:
                await update.message.reply_text("❌ Buyurtmalar yo‘q")
                return

            msg = "📦 So‘nggi buyurtmalar:\n\n"

            for r in rows:
                msg += f"🆔 {r[0]} | 💰 {r[1]} | 📌 {r[2]}\n"

            await update.message.reply_text(msg)

        elif text == "📊 Statistika":
            if update.effective_user.id != ADMIN_ID:
                return
            now = time.time()

            day = now - 86400
            week = now - 604800
            month = now - 2592000
            year = now - 31536000

            # 🔥 userlar
            cur.execute("SELECT COUNT(*) FROM users")
            total_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM users WHERE created_at >= %s", (day,))
            day_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM users WHERE created_at >= %s", (week,))
            week_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM users WHERE created_at >= %s", (month,))
            month_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM users WHERE created_at >= %s", (year,))
            year_users = cur.fetchone()[0]

            # 🔥 jami buyurtma
            cur.execute("SELECT COUNT(*) FROM orders")
            total_orders = cur.fetchone()[0]

            # 🔥 jami pul
            cur.execute("SELECT SUM(total) FROM orders")
            total_money = cur.fetchone()[0] or 0

            cur.execute("SELECT cart FROM orders")
            orders_data = cur.fetchall()

            total_profit = 0

            for row in orders_data:
                cart = json.loads(row[0])

                for pid, item in cart.items():
                    qty = item["qty"]

                    p = next((x for x in products if x["id"] == int(pid)), None)
                    if not p:
                        continue

                    price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                    cost = p.get("cost", 0)

                    # 🔥 SHU YERNI QO‘SH (ENG MUHIM)
                    if cost == 0:
                        continue

                    profit = (price - cost) * qty
                    total_profit += profit

            # 🔥 bugungi buyurtma
            cur.execute("""
            SELECT COUNT(*) FROM orders 
            WHERE DATE(to_timestamp(time)) = CURRENT_DATE
            """)
            today_orders = cur.fetchone()[0]

            # 🔥 bugungi pul
            cur.execute("""
            SELECT SUM(total) FROM orders 
            WHERE DATE(to_timestamp(time)) = CURRENT_DATE
            """)
            today_money = cur.fetchone()[0] or 0

            await update.message.reply_text(
                f"📊 STATISTIKA\n\n"

                f"👥 Jami user: {total_users}\n"
                f"📅 Bugun: {day_users}\n"
                f"📆 7 kun: {week_users}\n"
                f"🗓 1 oy: {month_users}\n"
                f"📈 1 yil: {year_users}\n\n"
                f"📈 FOYDA: {total_profit}\n\n"

                f"🧾 Jami buyurtma: {total_orders}\n"
                f"💰 Jami tushum: {total_money}\n\n"

                f"📦 Bugun buyurtma: {today_orders}\n"
                f"💵 Bugun tushum: {today_money}"
            )

        elif context.user_data.get("step") == "origin":
            origin = text.replace("🇺🇿 ", "").replace("🇨🇳 ", "").replace("🇹🇷 ", "").replace("🏭 ", "")
            context.user_data["origin"] = origin

            context.user_data["seasons"] = []
            context.user_data["step"] = "season"

            keyboard = [["☀️ Yozgi", "❄️ Qishki"], ["🌸 Bahor", "🍂 Kuz"]]

            await update.message.reply_text(
                "Fasl tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return    

        elif text == "✅ Tayyor" and context.user_data.get("step") == "season":
            context.user_data["step"] = "category"
        
            keyboard = get_category_buttons(context)
        
            await update.message.reply_text(
                "Kategoriya:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

        elif text == "ℹ️ Yordam":

            # 1-QISM
            await update.message.reply_text(
                "📏 1-QISM: Kiyimni qanday o‘lchash\n\n"
                "Bolaga mos eski kiyimni oling, stolga tekis qo‘ying va yuqoridan pastgacha uzunligini santimetrda o‘lchang. "
                "Masalan: 44 sm chiqdi. Shu o‘lcham eng muhim, chunki bot aynan shu bo‘yicha ishlaydi.\n\n"
                "Endi shu raqamni eslab qoling, chunki keyingi qadamda aynan shu orqali qidiruv qilasiz."
            )

            # 2-QISM
            await update.message.reply_text(
                "🔎 2-QISM: Qidirish va tanlash\n\n"
                "Botga 44 yozsangiz → 43, 44, 45 sm kiyimlar chiqadi. "
                "Bu sizga yaqin o‘lchamlarni ko‘rsatadi.\n\n"
                "Kattaroq kerak bo‘lsa → 46 yozing.\n"
                "Kichikroq kerak bo‘lsa → 42 yozing.\n\n"
                "Har bir kiyim ostida uzunligi yozilgan bo‘ladi, shu raqamga qarab tanlang."
            )

            # 3-QISM
            await update.message.reply_text(
                "🛒 3-QISM: Buyurtma berish\n\n"
                "Yoqgan kiyimni tanlab 🛒 Savatga qo‘shing.\n"
                "🧺 Savat ga kiring.\n"
                "🚚 Buyurtma berish ni bosing.\n"
                "Telefon raqamingizni yozing.\n\n"
                "Shu bilan buyurtma tugaydi va siz bilan bog‘lanishadi."
            )                   
        elif context.user_data.get("step") == "size_season" and text in ["☀️ Yozgi","❄️ Qishki","🌸 Bahor","🍂 Kuz"]:
            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "")
            context.user_data["filter_season"] = season
            context.user_data["step"] = "size_category"

        
            keyboard = get_category_buttons(context)
            await update.message.reply_text(
                "Kategoriya tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            ) 
            return  

        elif text == "❌ Lokatsiya ishlamayapti":
            user_id = update.effective_user.id
            cart = carts.get(user_id, {})

            if not cart:
                await update.message.reply_text("❌ Savat bo‘sh")
                return

            total = 0
            for pid, item in cart.items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                total += price * qty

            final = total + 0

            # 🔥 ENG MUHIM — TEMP ORDER
            context.user_data["temp_order"] = {
                "cart": cart,
                "location": {},   # 🔥 bo‘sh dict (None emas!)
                "total": final,
                "type": "delivery"
            }

            context.user_data["order_step"] = "phone"

            await update.message.reply_text(
        "📞 Telefon raqamingizni yozing:"
            )
        elif context.user_data.get("order_step") == "manual_location":
            address = text

            user_id = update.effective_user.id
            cart = carts.get(user_id, {})

            total = 0
            for pid, item in cart.items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                total += price * qty

            delivery = 0
            final = total + delivery

            context.user_data["temp_order"] = {
                "cart": cart,
                "location": {"text": address},
                "total": final,
                "type": "delivery"
            }

            context.user_data["order_step"] = "phone"

            await update.message.reply_text("📞 Telefon yuboring:")    

        elif text == "🔙 Orqaga":
            step = context.user_data.get("step")

        
            # 🔹 choose_type → gender
            if step == "choose_type":
                context.user_data["step"] = "user_gender"

                keyboard = [
                    ["👦 O‘g‘il", "👧 Qiz"],
                    ["🔙 Orqaga", "🏠 Bosh menyu"]
                ]

                await update.message.reply_text(
                    "Kim uchun:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )

            # 🔹 user_category → user_season
            elif step == "user_category":
                context.user_data["step"] = "user_season"

                keyboard = [
                    ["☀️ Yozgi","❄️ Qishki"],
                    ["🌸 Bahor","🍂 Kuz"],
                    ["🔙 Orqaga", "🏠 Bosh menyu"]
                ]

                await update.message.reply_text(
                    "Fasl tanlang:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )

            # 🔹 user_season → choose_type
            elif step == "user_season":
                context.user_data["step"] = "choose_type"

                keyboard = [
                    ["📂 Umumiy"],
                    ["🔙 Orqaga", "🏠 Bosh menyu"]
                ]

                await update.message.reply_text(
                    "Qanday qidirasiz?",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )

            # 🔹 default → bosh menyu
            else:
                context.user_data.clear()
                await update.message.reply_text("🏠 Bosh menyu", reply_markup=MAIN_MENU)

            return
        
        elif context.user_data.get("step") == "season":
            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "")
        
            if "seasons" not in context.user_data:
                context.user_data["seasons"] = []
        
            if season not in context.user_data["seasons"]:
                context.user_data["seasons"].append(season)
        
            keyboard = [
                ["☀️ Yozgi","❄️ Qishki"],
                ["🌸 Bahor","🍂 Kuz"],
                ["✅ Tayyor"]
            ]
        
            await update.message.reply_text(
                f"Tanlangan: {', '.join(context.user_data['seasons'])}",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        
        elif context.user_data.get("step") == "category":

            category = text.split("(")[0]
            category = category.replace("👕","").replace("👖","").replace("🧥","") \
                            .replace("🩳","").replace("👟","").replace("🧢","").replace("🩲","").strip().lower()

            # 🔥 TO‘G‘RI NOMGA O‘TKAZAMIZ
            if "2 talik" in category:
                category = "2 talik kiyim"
            elif "3 talik" in category:
                category = "3 talik kiyim"
            elif "futbolka" in category:
                category = "futbolka"
            elif "shim" in category:
                category = "shim"
            elif "qalin" in category:
                category = "qalin kiyim"
            elif "shortik" in category:
                category = "shortik"
            elif "oyoq" in category:
                category = "oyoq kiyim"
            elif "bosh" in category:
                category = "bosh kiyim"
            elif "ichki" in category:
                category = "ichki kiyim"

            context.user_data["category"] = category.strip().lower()
            context.user_data["step"] = "name"

            await update.message.reply_text("Nomini yozing:")
            return

        elif context.user_data.get("step") == "all_season":

            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "").strip().lower()

            found = False

            for p in products:
                try:
                    p_seasons = p.get("season", [])

                    if not isinstance(p_seasons, list):
                        p_seasons = str(p_seasons).split(",")

                    p_seasons = [str(s).strip().lower() for s in p_seasons]

                    if season in p_seasons:

                        photo = p.get("photo")
                        if not photo:
                            continue

                        found = True

                        await update.message.reply_photo(
                            photo=photo,
                            caption=f"{p.get('name','')}\n📏 {p.get('size','')}\n💰 {p.get('price','')}"
                        )

                except Exception as e:
                    print("XATO:", p, e)
                    continue

            if not found:
                await update.message.reply_text("❌ Mahsulot yo‘q")

            context.user_data.clear()


        elif context.user_data.get("step") == "name":
            context.user_data["name"] = text
            context.user_data["step"] = "size"

            await update.message.reply_text("📏 Uzunlik yozing (sm) (masalan 40):")
            return
        elif context.user_data.get("step") == "size":
            size = text.strip()

            if not size.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing (masalan 44)")
                return

            context.user_data["size"] = size
            context.user_data["step"] = "price"

            await update.message.reply_text("Narxni yozing:")
                    
        elif context.user_data.get("step") == "price":
            price = text.replace(" ", "").replace("so'm","").replace("soʻm","")

            if not price.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing (masalan 50000)")
                return

            price = int(price)
            price = f"{price:,}".replace(",", " ")

            context.user_data["price"] = price + " so‘m"
            context.user_data["step"] = "cost"

            await update.message.reply_text("💸 Tannarxni yozing (masalan 30000):")

        elif context.user_data.get("step") == "cost":
            cost = text.replace(" ", "")

            if not cost.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing")
                return

            context.user_data["cost"] = int(cost)
            context.user_data["step"] = "count"

            await update.message.reply_text("📦 Nechta bor? (masalan 4):")

        elif context.user_data.get("step") == "count":
            if not text.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing")
                return

            context.user_data["count"] = int(text)

            # 🔥 xavfsiz olish
            photo = context.user_data.get("photo")
            gender = context.user_data.get("gender")
            origin = context.user_data.get("origin")
            seasons = context.user_data.get("seasons", [])
            category = context.user_data.get("category")
            name = context.user_data.get("name")
            size = context.user_data.get("size")
            price = context.user_data.get("price")
            cost = context.user_data.get("cost", 0)
            count = context.user_data.get("count")

            if context.user_data.get("mode") == "edit":
                pid = context.user_data.get("edit_product_id")

                cur.execute("""
                UPDATE products
                SET photo=%s, gender=%s, origin=%s, season=%s,
                    category=%s, name=%s, size=%s, price=%s, count=%s
                WHERE id=%s
                """, (
                    photo, gender, origin,
                    ",".join(seasons),
                    category, name, size, price, count,
                    pid
                ))

                await update.message.reply_text("✅ Tahrirlandi!")
            else:
                cur.execute("""
                INSERT INTO products (
                    photo, gender, origin, season, category, name, size, price, count, reserved, cost
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    photo,
                    gender,
                    origin,
                    ",".join(seasons),
                    category,
                    name,
                    size,
                    price,
                    count,
                    0,
                    context.user_data.get("cost", 0)
                ))

                await update.message.reply_text("✅ Qo‘shildi!")

            conn.commit()
            load_products_from_db()
            context.user_data.clear()
        elif context.user_data.get("step") == "size_filter":
            size = text.replace(" ", "")
            context.user_data["filter_size"] = size
            context.user_data["step"] = "size_season"

            keyboard = [
                ["☀️ Yozgi","❄️ Qishki"],
                ["🌸 Bahor","🍂 Kuz"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Fasl tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

            # ===== USER FLOW =====
        elif text == "🛍 Kiyimlar":
            context.user_data.clear() 
            keyboard = [["👦 O‘g‘il", "👧 Qiz"],["🔙 Orqaga", "🏠 Bosh menyu"]]
            await update.message.reply_text("Tanlang:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))

        # 👦 / 👧
        elif text in ["👦 O‘g‘il", "👧 Qiz"] and context.user_data.get("step") != "gender":
            gender = text.replace("👦 ", "").replace("👧 ", "")
            context.user_data["filter_gender"] = gender

            context.user_data["step"] = "origin_select"

            keyboard = [
                ["🇺🇿 Vodiy", "🇨🇳 Xitoy"],
                ["🇹🇷 Turkiya", "🏭 8-mart fabrika"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Qaysi ishlab chiqarish:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

        elif text == "📂 Umumiy":

            context.user_data["step"] = "user_season"

            keyboard = [
                ["☀️ Yozgi","❄️ Qishki"],
                ["🌸 Bahor","🍂 Kuz"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Fasl tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )


        elif text == "🧺 Savat":
            load_products_from_db()
            user_id = update.effective_user.id
            cart = clean_cart(user_id, context)

            
            now = time.time()
            new_cart = {}

            for pid, item in cart.items():
                if now - item["time"] < 7200:
                    new_cart[pid] = item
                else:
                    qty = item["qty"]
                    p = next((x for x in products if x["id"] == int(pid)), None)
                    if p:
                        p["reserved"] = max(0, p.get("reserved", 0) - qty)

            carts[user_id] = new_cart
            cart = new_cart

            if not cart:
                await update.message.reply_text("🧺 Savat bo‘sh")
                return

            msg = "🧺 Savat:\n\n"
            total = 0
            keyboard = []

            for pid, item in cart.items():
                qty = item["qty"]

                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

                summa = price * qty
                total += summa

                msg += f"{p['name']} x{qty} = {summa}\n"

                keyboard.append([
                    InlineKeyboardButton("❌", callback_data=f"del_{pid}")
                ])

            msg += f"\n💰 Jami: {total}"

            keyboard.append([InlineKeyboardButton("🚚 Buyurtma", callback_data="checkout")])
            keyboard.append([InlineKeyboardButton("🔙 Orqaga", callback_data="back")])

            await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard))

        # 🌦 FASL
        elif text in ["☀️ Yozgi","❄️ Qishki","🌸 Bahor","🍂 Kuz"]:
            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "")
            context.user_data["filter_season"] = season
            context.user_data["step"] = "user_category"

            keyboard = get_category_buttons(context)
            await update.message.reply_text(
                "Kategoriya tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        elif context.user_data.get("step") == "origin_select":
            origin = text.replace("🇺🇿 ", "").replace("🇨🇳 ", "").replace("🇹🇷 ", "").replace("🏭 ", "")
            context.user_data["filter_origin"] = origin

            context.user_data["step"] = "choose_type"

            keyboard = [
                ["📂 Umumiy"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Qanday qidirasiz?",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return


        elif "(" in text and context.user_data.get("step") == "user_category":

            category = text.split("(")[0]
            category = category.replace("👕","").replace("👖","").replace("🧥","") \
                            .replace("🩳","").replace("👟","").replace("🧢","").replace("🩲","").strip().lower()

            if "2 talik" in category:
                category = "2 talik kiyim"
            elif "3 talik" in category:
                category = "3 talik kiyim"
            elif "futbolka" in category:
                category = "futbolka"
            elif "shim" in category:
                category = "shim"
            elif "qalin" in category:
                category = "qalin kiyim"
            elif "shortik" in category:
                category = "shortik"
            elif "oyoq" in category:
                category = "oyoq kiyim"
            elif "bosh" in category:
                category = "bosh kiyim"
            elif "ichki" in category:
                category = "ichki kiyim"

            context.user_data["filter_category"] = category

            # 🔥 FILTER
            filtered = []

            for p in products:

                # kategoriya
                if category not in str(p.get("category", "")).lower():
                    continue

                # origin
                origin = context.user_data.get("filter_origin")
                if origin:
                    if origin.lower() not in str(p.get("origin", "")).lower():
                        continue

                # season
                season = context.user_data.get("filter_season")
                if season:
                    p_seasons = p.get("season", [])

                    if not isinstance(p_seasons, list):
                        p_seasons = str(p_seasons).split(",")

                    p_seasons = [s.strip().lower() for s in p_seasons]

                    if season.lower() not in p_seasons:
                        continue

                # mavjudlik
                available = p.get("count", 0) - p.get("reserved", 0)
                if available <= 0:
                    continue

                filtered.append(p)

            # ❗ BU YERDA ENDI FILTERNI TEKSHIRAMIZ
            if not filtered:
                await update.message.reply_text("❌ Mos mahsulot topilmadi")
                return

            context.user_data["filtered"] = filtered
            context.user_data["i"] = 0

            p = filtered[0]
            photo = p.get("photo")

            keyboard = [
                [
                    InlineKeyboardButton("⬅️", callback_data="prev_one"),
                    InlineKeyboardButton("➡️", callback_data="next_one")
                ],
                [InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p.get('id')}")]
            ]

            # 🔥 RASM BOR/YO‘Q HOLAT
            if photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=f"1/{len(filtered)}\n\n{p.get('name')}\n📏 {p.get('size')}\n💰 {p.get('price')}",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            else:
                await update.message.reply_text(
                    f"1/{len(filtered)}\n\n{p.get('name')}\n📏 {p.get('size')}\n💰 {p.get('price')}\n\n⚠️ Rasm yo‘q",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
        elif text == "🚚 Buyurtma berish":
            keyboard = [["🚚 Dastavka", "📍 Olib ketish"],["🔙 Orqaga", "🏠 Bosh menyu"]]

            await update.message.reply_text(
                "Qanday olasiz?",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )   
        elif text == "🚚 Dastavka":
            context.user_data["order_step"] = "location"
            context.user_data["order_type"] = "delivery"

            keyboard = [[KeyboardButton("📍 Lokatsiya yuborish", request_location=True)],
            ["❌ Lokatsiya ishlamayapti"],
            ["🏠 Bosh menyu"]
        ]

            await update.message.reply_text(
                " Dastavka narxi taxminan 20 000-50 000 so‘m atrofida bo‘ladi\n 📍 Lokatsiyangizni yuboring va \n⏳ Iltimos bir oz kuting... yoki lokatsiya ishlamasa pastdagi tugmani bosing:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )  

        elif context.user_data.get("order_step") == "phone":
            phone = text.strip().replace(" ", "")

            if phone.isdigit() and len(phone) == 9:
                phone = "+998" + phone
            elif phone.startswith("+998") and len(phone) == 13:
                pass
            else:
                await update.message.reply_text("❌ Noto‘g‘ri raqam!")
                return

            data = context.user_data.get("temp_order")
            if not data:
                await update.message.reply_text("❌ Xatolik")
                return

            user_id = update.effective_user.id
            cur.execute("""
            INSERT INTO orders (user_id, cart, location, phone, total, status, time)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
            """, (
                user_id,
                json.dumps(data["cart"]),
                json.dumps(data["location"]),
                phone,
                data["total"],
                "new",
                time.time()
            ))

            order_id = str(cur.fetchone()[0])
            conn.commit()

            # ===== MAHSULOTNI KAMAYTIRISH =====
            for pid, item in data["cart"].items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if p:
                    p["count"] -= qty
                    p["reserved"] = max(0, p.get("reserved", 0) - qty)
                    # 🔥 DB ga ham yozamiz
                    cur.execute(
                        "UPDATE products SET count = count - %s, reserved = GREATEST(0, reserved - %s) WHERE id = %s",
                        (qty, qty, p["id"])
                    )
            conn.commit()
            #save_products()

            # ===== USERGA MAHSULOT =====
            for pid, item in data["cart"].items():
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue
                qty = item["qty"]

                await context.bot.send_photo(
                    chat_id=user_id,
                    photo=p["photo"],
                    caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
                )

            # ===== USER STATUS =====
            if data.get("type") == "delivery":
                await update.message.reply_text(
                    "🚚 Buyurtma qabul qilindi",
                    reply_markup=MAIN_MENU
                )
            else:
                await update.message.reply_text(
                    "📍 Olib ketish manzili:\nSamarqand, Pastdarg‘om, Charxin\nA'loqa 📞 +998915388499  Adminlar o'zlari a'loqaga chiqishadi va manzilni yetgazishadi. " 
                    ,
                    reply_markup=MAIN_MENU
                )

                #await context.bot.send_location(
                #   chat_id=user_id,
                #  latitude=39.690149,
                # longitude=66.824828
                #)

                await update.message.reply_text(
                    "🏠 Bosh menyu",
                    reply_markup=MAIN_MENU
                )

            # ===== ADMIN TUGMALAR =====
            admin_keyboard = [
                [InlineKeyboardButton("📞 Aloqa", callback_data=f"contact_{order_id}")],
                [InlineKeyboardButton("🚚 Yetkazishni boshlash", callback_data=f"deliver_{order_id}")],
                [InlineKeyboardButton("✅ Yakunlandi", callback_data=f"done_{order_id}")],
                [InlineKeyboardButton("❌ Bekor", callback_data=f"cancel_{order_id}")]
            ]

            # ===== ADMINGA MAHSULOT =====
            for pid, item in data["cart"].items():
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue
                qty = item["qty"]

                await context.bot.send_photo(
                    chat_id=ADMIN_ID,
                    photo=p["photo"],
                    caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
                )

            # ===== ADMINGA UMUMIY INFO =====
            if data.get("type") == "delivery":
                text_admin = f"🚚 DASTAVKA\n📞 {phone}\n💰 {data['total']}"
            else:
                text_admin = f"📍 OLIB KETISH\n📞 {phone}\n💰 {data['total']}"

            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=text_admin,
                reply_markup=InlineKeyboardMarkup(admin_keyboard)
            )

            # ===== TOZALASH =====
            carts[user_id] = {}
            context.user_data.clear()

        elif text == "📍 Olib ketish":
            context.user_data["order_step"] = "phone"

            user_id = update.effective_user.id
            cart = carts.get(user_id, {})

            if not cart:
                await update.message.reply_text("❌ Savat bo‘sh")
                return

            total = 0
            for pid, item in cart.items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                total += price * qty

            context.user_data["temp_order"] = {
                "cart": cart,
                "location": None,
                "total": total,
                "type": "pickup"
            }

            keyboard = [
                [KeyboardButton("📞 Telefon yuborish", request_contact=True)],
                ["🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "📞 Telefon raqamingizni yuboring yoki yozing:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

        elif context.user_data.get("step") == "inline_category" and "(" in text:

            category = text.split("(")[0]
            category = category.replace("👕","").replace("👖","").replace("🧥","") \
                .replace("🩳","").replace("👟","").replace("🧢","").replace("🩲","").strip().lower()

            # 🔥 to‘g‘ri nomga o‘tkazamiz
            if "2 talik" in category:
                category = "2 talik kiyim"
            elif "3 talik" in category:
                category = "3 talik kiyim"
            elif "futbolka" in category:
                category = "futbolka"
            elif "shim" in category:
                category = "shim"
            elif "qalin" in category:
                category = "qalin kiyim"
            elif "shortik" in category:
                category = "shortik"
            elif "oyoq" in category:
                category = "oyoq kiyim"
            elif "bosh" in category:
                category = "bosh kiyim"
            elif "ichki" in category:
                category = "ichki kiyim"

            context.user_data["filter_category"] = category
            context.user_data["step"] = "write_size"

            await update.message.reply_text(
                "📏 Mahsulot uzunligini yozing (sm)\nMasalan: 44\n Bu kiyimdagi o‘lcham (razmer) emas maxsulotning uzunligi. Tushunmasangiz yordam bo‘limiga o‘ting bosh menyudan "
            )
        elif context.user_data.get("step") == "write_size":

            size = text.strip()

            if not size.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing (masalan 44)")
                return

            context.user_data["filter_size"] = size

            found = False

            for p in products:
                if filter_check(p, context):
                    found = True

                    keyboard = [
                        [InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p['id']}")]
                    ]

                    await update.message.reply_photo(
                        photo=p["photo"],
                        caption=f"{p['name']}\n📏 {p['size']} sm\n💰 {p['price']}",
                        reply_markup=InlineKeyboardMarkup(keyboard)
                    )

            # ❗ AGAR TOPILMASA
            if not found:
                await update.message.reply_text(
                    "❌ Mos mahsulot topilmadi.\n\nBoshqa razmer yozing (masalan 42, 46)"
                )
                return  # 🔥 MUHIM — step o‘chmaydi

            # 🔥 FAqat topilganda tozalaymiz
            context.user_data.clear()
        elif text == "🏠 Bosh menyu":
            context.user_data.clear()

            await update.message.reply_text(
                "🏠 Bosh menyu",
                reply_markup=MAIN_MENU
            )
            return

    except Exception as e:
        print("XATO:", e)
        await update.message.reply_text("❌ Xatolik yuz berdi")
        context.user_data.clear()

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    data = query.data


# ===== FILTER BLOK =====
    if data.startswith("g_") or data.startswith("o_") or data.startswith("s_"):

        if data.startswith("g_"):
            g = data[2:]

            if "o‘g" in g or "og" in g:
                context.user_data["filter_gender"] = "o‘g‘il"
            else:
                context.user_data["filter_gender"] = "qiz"

        elif data.startswith("o_"):
            context.user_data["filter_origin"] = data[2:]

        elif data.startswith("s_"):
            context.user_data["filter_season"] = data[2:]

        # 🔥 TEXTNI YANGILAYMIZ
        await query.message.edit_text(
            f"🔎 Tanlang:\n\n"
            f"Jins: {context.user_data.get('filter_gender','-')}\n"
            f"Fabrika: {context.user_data.get('filter_origin','-')}\n"
            f"Fasl: {context.user_data.get('filter_season','-')}\n",
            reply_markup=get_filter_menu(context.user_data)
        )
        return
    # ===== RESET =====
    if data == "reset":
        context.user_data.clear()

        await query.message.edit_text(
            "🔎 Tanlang:\n\nJins: -\nFabrika: -\nFasl: -\nRazmer: -",
            reply_markup=get_filter_menu(context.user_data)
        )
        return


    # ===== APPLY =====
    if data == "apply":

        context.user_data["step"] = "inline_category"

        keyboard = get_category_buttons(context)

        await query.message.reply_text(
            "📂 Kategoriya tanlang:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    elif data.startswith("add_"):
        
        product_id = int(data.split("_")[1])

        if product_id not in product_locks:
            product_locks[product_id] = asyncio.Lock()

        async with product_locks[product_id]:

            product = next((x for x in products if x["id"] == product_id), None)
            if not product:
                await query.answer("❌ Topilmadi", show_alert=True)
                return

            # 🔥 REAL TEKSHIRUV
            available = product["count"] - product.get("reserved", 0)

            if available <= 0:
                await query.answer("❌ Bu mahsulot band yoki qolmagan", show_alert=True)
                return

            if user_id not in carts:
                carts[user_id] = {}

            if product_id in carts[user_id]:
                carts[user_id][product_id]["qty"] += 1
            else:
                carts[user_id][product_id] = {
                    "qty": 1,
                    "time": time.time()
                }

            carts[user_id][product_id]["time"] = time.time()

            # 🔥 ENG MUHIM — RESERVE
            product["reserved"] = product.get("reserved", 0) + 1

        await query.answer("✅ Savatga qo‘shildi")

        keyboard = [
            [InlineKeyboardButton("🧺 Savatga o‘tish", callback_data="go_cart")]
        ]

        await query.message.reply_text(
            "🛒 Mahsulot vaqtincha siz uchun band qilindi!\n⏳ 2 soat ichida xarid qilmasangiz o‘chiriladi.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data.startswith("edit_"):
        product_id = int(data.split("_")[1])

        old_product = next((x for x in products if x["id"] == product_id), None)
        if not old_product:
            await query.message.reply_text("❌ Topilmadi")
            return

        # eski mahsulotni o‘chir
        cur.execute("DELETE FROM products WHERE id=%s", (product_id,))
        conn.commit()
        load_products_from_db()

        # bot rasmni chiqaradi
        await context.bot.send_photo(
            chat_id=query.message.chat_id,
            photo=old_product["photo"]
        )

        # 🔥 ENG MUHIM — bot yuborgan rasmni ham ishlatamiz
        await start_photo_flow(context, query.message.chat_id, old_product["photo"])
    elif data.startswith("delete_"):
        if query.from_user.id != ADMIN_ID:
            return

        product_id = int(data.split("_")[1])

        cur.execute("DELETE FROM products WHERE id=%s", (product_id,))
        conn.commit()

        load_products_from_db()

        await query.message.reply_text("✅ Mahsulot o‘chirildi")
    elif data == "clear_yes":
        if query.from_user.id != ADMIN_ID:
            return

        cur.execute("DELETE FROM products")
        conn.commit()
        load_products_from_db()
       # save_products()

        await query.message.reply_text("✅ Barcha mahsulotlar o‘chirildi")

    elif data == "clear_no":
        await query.message.reply_text("❌ Bekor qilindi")

            # 🔥 oldinga orqaga
    elif data == "next_one" or data == "prev_one":

        # 🔄 indexni o‘zgartiramiz
        if data == "next_one":
            context.user_data["i"] += 1
            if context.user_data["i"] >= len(context.user_data["filtered"]):
                context.user_data["i"] = len(context.user_data["filtered"]) - 1

        else:  # prev_one
            context.user_data["i"] -= 1
            if context.user_data["i"] < 0:
                context.user_data["i"] = 0

        # 🔥 MAHSULOTNI OLAMIZ
        p = context.user_data["filtered"][context.user_data["i"]]

        # 🔘 TUGMALAR
        keyboard = []
        nav = []

        if context.user_data["i"] > 0:
            nav.append(InlineKeyboardButton("⬅️", callback_data="prev_one"))

        if context.user_data["i"] < len(context.user_data["filtered"]) - 1:
            nav.append(InlineKeyboardButton("➡️", callback_data="next_one"))

        if nav:
            keyboard.append(nav)

        keyboard.append([
            InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p.get('id')}")
        ])

        if update.effective_user.id == ADMIN_ID:
            keyboard.append([
                InlineKeyboardButton("✏️ Edit", callback_data=f"edit_{p.get('id')}"),
                InlineKeyboardButton("🗑 O‘chirish", callback_data=f"delete_{p.get('id')}")
            ])

        from telegram import InputMediaPhoto

        await query.message.edit_media(
            media=InputMediaPhoto(
                media=p.get("photo"),
                caption=f"{context.user_data['i']+1}/{len(context.user_data['filtered'])}\n\n"
                        f"{p.get('name')}\n📏 {p.get('size')}\n💰 {p.get('price')}"
            ),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data.startswith("plus_"):
        product_id = int(data.split("_")[1])

        product = next((x for x in products if x["id"] == product_id), None)
        if not product:
            return

        if product["count"] - product.get("reserved", 0) <= 0:
            await query.answer("❌ Yetarli mahsulot yo‘q", show_alert=True)
            return

        if user_id not in carts:
            carts[user_id] = {}

        # 🔥 ENG MUHIM TUZATISH
        if product_id in carts[user_id]:
            carts[user_id][product_id]["qty"] += 1
        else:
            carts[user_id][product_id] = {
                "qty": 1,
                "time": time.time()
            }

        carts[user_id][product_id]["time"] = time.time()
        product["reserved"] += 1

        await query.answer("➕ Qo‘shildi")
    elif data.startswith("send_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USERGA
        await context.bot.send_message(
            chat_id=user_id,
            text="🚚 Buyurtmangiz yo‘lga chiqdi!\n⏳ 1 soat ichida yetkaziladi."
        )

        # ADMINGA
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"🚚 Buyurtma jo‘natildi\nID: {order_id}"
        )

        await query.answer("Yuborildi")

 
    elif data.startswith("deliver_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USERGA
        await context.bot.send_message(
            chat_id=user_id,
            text="🚚 Buyurtmangiz yetkazilmoqda!\n⏳ 1 soat ichida yetkaziladi."
        )

        # ADMINGA
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"🚚 Yetkazish boshlandi\nID: {order_id}"
        )

        await query.answer("Yuborildi")

    elif data.startswith("done_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="✅ Buyurtmangizni qabul qilib oldingiz. Rahmat 😊"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"✅ YAKUNLANDI\nID: {order_id}"
        )

        cur.execute("DELETE FROM orders WHERE id=%s", (order_id,))
        conn.commit()
       # save_orders()

        await query.answer("Yakunlandi")

    elif data.startswith("ready_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📦 Buyurtmangiz tayyor!"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 TAYYOR\nID: {order_id}"
        )

        await query.answer("Tayyor qilindi")

    elif data.startswith("cancel_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id, cart FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]
        cart = json.loads(row[1])

        # 🔥 mahsulotni qaytaramiz
        for product_id, item in cart.items():
            qty = item["qty"]

            p = next((x for x in products if x["id"] == int(product_id)), None)

            if p:
                p["count"] += qty
                p["reserved"] = max(0, p["reserved"] - qty)

        load_products_from_db()  # 🔥 ENG MUHIM

        carts[user_id] = {}

        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Buyurtmangiz bekor qilindi"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"❌ BUYURTMA BEKOR QILINDI\nID: {order_id}"
        )

        cur.execute("DELETE FROM orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("Bekor qilindi")
    elif data.startswith("user_cancel_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id, cart FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            await query.answer("Allaqachon bekor qilingan")
            return

        user_id = row[0]
        cart = json.loads(row[1])

        # 🔥 mahsulotni qaytaramiz
        for pid, item in cart.items():
            qty = item["qty"]
            p = next((x for x in products if x["id"] == int(pid)), None)
            if p:
                p["count"] += qty
                p["reserved"] = max(0, p.get("reserved", 0) - qty)

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"❌ Mijoz bekor qildi\nID: {order_id}"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Buyurtma bekor qilindi"
        )

        cur.execute("DELETE FROM orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("Bekor qilindi")

    elif data.startswith("delivered_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USER ga
        await context.bot.send_message(
            chat_id=user_id,
            text=f"📦 Buyurtma yetkazildi!\n🆔 ID: {order_id}"
        )

        # ADMIN ga
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 Yetkazildi\nID: {order_id}"
        )

        await query.answer("Yetkazildi")
    elif data.startswith("paid_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USER ga
        await context.bot.send_message(
            chat_id=user_id,
            text="💰 To‘lov qabul qilindi! Rahmat 😊"
        )

        # ADMIN ga
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"💰 To‘lov olindi\nID: {order_id}"
        )

        # ORDERNI O‘CHIRAMIZ
        cur.execute("DELETE FROM orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("To‘lov olindi")

    elif data.startswith("accept_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # 🔥 ADMIN MA’LUMOTI
        ADMIN_PHONE = "+998915388499"
        ADDRESS = "Samarqand vil. Pastdarg'om tum. charxin shax. charos ko‘chasi 53 uy Adminlar aloqaga chiqishadi va olib ketish vaqtini keishiladi."

        #LAT = 39.690149
        #LON = 66.824828

        # USERGA
        await context.bot.send_message(
            chat_id=user_id,
            text=f"✅ Buyurtmangiz qabul qilindi!\n\n📞 Tel: {ADMIN_PHONE}\n🏠 Manzil: {ADDRESS}"
        )

      #  await context.bot.send_location(
       #     chat_id=user_id,
        #    latitude=LAT,
         #   longitude=LON
        #)

        # ADMIN ga
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"✅ BUYURTMA QABUL QILINDI\nID: {order_id}"
        )

        await query.answer("Yuborildi")
    elif data.startswith("contact_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📞 Admin sizga 10 min ichida bog‘lanadi"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📞 Mijoz bilan bog‘laning\nID: {order_id}"
        )

        await query.answer("Yuborildi")
    elif data.startswith("picked_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📦 Buyurtma yakunlandi! Rahmat 😊"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 BUYURTMA YAKUNLANDI\nID: {order_id}"
        )

        # 🔥 ORDERNI O‘CHIRAMIZ
        cur.execute("DELETE FROM orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("Yakunlandi")

    elif data == "go_cart":
        user_id = query.from_user.id
        cart = carts.get(user_id, {})

        
        now = time.time()
        new_cart = {}

        # ⏳ 2 soatdan eski itemlarni tozalash
        for pid, item in cart.items():
            if now - item["time"] < 7200:
                new_cart[pid] = item
            else:
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if p:
                    p["reserved"] = max(0, p.get("reserved", 0) - qty)

        carts[user_id] = new_cart
        cart = new_cart

        # 🧺 Savat bo‘sh bo‘lsa
        if not cart:
            await query.message.reply_text("🧺 Savat bo‘sh")
            return

        msg = "🧺 Savat:\n\n"
        total = 0
        keyboard = []

        # 📦 Savat ichidagi mahsulotlar
        for pid, item in cart.items():
            try:
                pid = int(pid)
            except:
                continue

            p = next((x for x in products if x["id"] == pid), None)
            if not p:
                continue

            qty = item["qty"]

            # 💰 narxni tozalab olish
            price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

            summa = price * qty
            total += summa

            msg += f"{p['name']} x{qty} = {summa}\n"

            keyboard.append([
                InlineKeyboardButton(f"❌ {p['name']}", callback_data=f"del_{pid}")
            ])

        msg += f"\n💰 Jami: {total}"

        keyboard.append([
            InlineKeyboardButton("🚚 Buyurtma berish", callback_data="checkout")
        ])
        keyboard.append([
            InlineKeyboardButton("🔙 Orqaga", callback_data="back")
        ])

        await query.message.reply_text(
            msg,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    elif data.startswith("del_"):
        product_id = int(data.split("_")[1])

        cart = carts.get(user_id, {})

        # 🔥 cart kaliti int yoki str bo'lishi mumkin — ikkalasini tekshiramiz
        cart_key = product_id if product_id in cart else str(product_id) if str(product_id) in cart else None

        if cart_key is not None:
            qty = cart[cart_key]["qty"]

            p = next((x for x in products if x["id"] == product_id), None)
            if p:
                p["reserved"] = max(0, p.get("reserved", 0) - qty)

            cart.pop(cart_key)

        await query.answer("❌ O‘chirildi")

        if not cart:
            await query.message.reply_text("🧺 Savat bo‘sh")
            return

        msg = "🧺 Savat:\n\n"
        total = 0
        keyboard = []

        for pid, item in cart.items():
            qty = item["qty"]

            p = next((x for x in products if x["id"] == int(pid)), None)
            if not p:
                continue

            price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

            summa = price * qty
            total += summa

            msg += f"{p['name']} x{qty} = {summa}\n"

            keyboard.append([
                InlineKeyboardButton("❌", callback_data=f"del_{pid}")
            ])

        msg += f"\n💰 Jami: {total}"

        keyboard.append([InlineKeyboardButton("🚚 Buyurtma berish", callback_data="checkout")])
        keyboard.append([InlineKeyboardButton("🔙 Orqaga", callback_data="back")])

        await query.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard))
    elif data == "checkout":

        context.user_data["order_started"] = time.time()
        context.user_data["order_step"] = "choose_type"

        keyboard = [
            ["🚚 Dastavka", "📍 Olib ketish"],
            ["🏠 Bosh menyu"]
        ]

        await query.message.reply_text(
            "Qanday olasiz?",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )


    elif data == "back":
        await query.message.reply_text(
            "🏠 Bosh menyu",
            reply_markup=MAIN_MENU
            
        )
    elif data.startswith("confirm_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📦 Buyurtmangiz tayyor! Kelishilgan vaqtda olib ketishingiz mumkin"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 BUYURTMA TASDIQLANDI\nID: {order_id}"
        )

        await query.answer("Tasdiqlandi")
async def location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if context.user_data.get("order_step") != "location":
        return

    location = update.message.location
    lat = location.latitude
    lon = location.longitude

    user_id = update.effective_user.id
    cart = carts.get(user_id, {})

    if not cart:
        await update.message.reply_text("❌ Savat bo‘sh")
        return

    total = 0

    for pid, item in cart.items():
        qty = item["qty"]
        p = next((x for x in products if x["id"] == int(pid)), None)
        if not p:
            continue

        price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

        total += price * qty   # ✅ TO‘G‘RI

    delivery = 0
    final = total + delivery

    context.user_data["temp_order"] = {
        "cart": cart,
        "location": {"lat": lat, "lon": lon},
        "total": final,
        "type": context.user_data.get("order_type")
    }

    await update.message.reply_text("⏳ Lokatsiya qabul qilindi, hisoblanmoqda...")

    context.user_data["order_step"] = "phone"

    keyboard = [
        [KeyboardButton("📞 Telefon yuborish", request_contact=True)],
        ["🏠 Bosh menyu"]
    ]

    await update.message.reply_text(
        "📞 Telefon raqamingizni yuboring yoki yozing (+998...):",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("order_step") != "phone":
        return

    contact = update.message.contact
    phone = contact.phone_number

    if not phone.startswith("+"):
        phone = "+" + phone

    data = context.user_data.get("temp_order")
    if not data:
        await update.message.reply_text("❌ Xatolik")
        return

    user_id = update.effective_user.id
    cur.execute("""
    INSERT INTO orders (user_id, cart, location, phone, total, status, time)
    VALUES (%s,%s,%s,%s,%s,%s,%s)
    RETURNING id
    """, (
        user_id,
        json.dumps(data["cart"]),
        json.dumps(data["location"]),
        phone,
        data["total"],
        "new",
        time.time()
    ))

    order_id = str(cur.fetchone()[0])
    conn.commit()

    # ===== MAHSULOTNI KAMAYTIRISH =====
    for pid, item in data["cart"].items():
        qty = item["qty"]
        p = next((x for x in products if x["id"] == int(pid)), None)
        if p:
            p["count"] -= qty
            p["reserved"] = max(0, p.get("reserved", 0) - qty)
            # 🔥 DB ga ham yozamiz
            cur.execute(
                "UPDATE products SET count = count - %s, reserved = GREATEST(0, reserved - %s) WHERE id = %s",
                (qty, qty, p["id"])
            )
    conn.commit()

    # ===== USERGA MAHSULOT =====
    for pid, item in data["cart"].items():
        p = next((x for x in products if x["id"] == int(pid)), None)
        if not p:
            continue
        qty = item["qty"]

        await context.bot.send_photo(
            chat_id=user_id,
            photo=p["photo"],
            caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
        )

    # ===== ADMINGA MAHSULOT =====
    for pid, item in data["cart"].items():
        p = next((x for x in products if x["id"] == int(pid)), None)
        if not p:
            continue
        qty = item["qty"]

        await context.bot.send_photo(
            chat_id=ADMIN_ID,
            photo=p["photo"],
            caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
        )

    # ===== ADMIN TUGMALAR =====
    if data.get("type") == "delivery":
        admin_keyboard = [
            [InlineKeyboardButton("📞 Aloqa", callback_data=f"contact_{order_id}")],
            [InlineKeyboardButton("🚚 Buyurtmani jo'natish", callback_data=f"send_{order_id}")],
            [InlineKeyboardButton("✅ Yakunlandi", callback_data=f"done_{order_id}")],
            [InlineKeyboardButton("❌ Bekor", callback_data=f"cancel_{order_id}")]
        ]
    else:
        admin_keyboard = [
            [InlineKeyboardButton("📞 Aloqa", callback_data=f"contact_{order_id}")],
            [InlineKeyboardButton("📦 Buyurtmani tasdiqlash", callback_data=f"confirm_{order_id}")],
            [InlineKeyboardButton("✅ Yakunlandi", callback_data=f"done_{order_id}")],
            [InlineKeyboardButton("❌ Bekor", callback_data=f"cancel_{order_id}")]
        ]

    # ===== TEXT =====
    if data.get("type") == "delivery":
        if data.get("location") and "lat" in data["location"]:
            lat = data["location"]["lat"]
            lon = data["location"]["lon"]
            loc = f"\n📍 https://maps.google.com/?q={lat},{lon}"
        else:
            loc = "\n📍 Lokatsiya yuborilmadi"

        text_admin = (
            f"🚚 DASTAVKA\n"
            f"📞 {phone}\n"
            f"💰 {data['total']}{loc}"
        )

        await update.message.reply_text(
            "🚚 Buyurtma qabul qilindi!\n📞 Admin siz bilan tez orada bog'lanadi.",
            reply_markup=MAIN_MENU
        )
    else:
        text_admin = (
            f"📍 OLIB KETISH\n"
            f"📞 {phone}\n"
            f"💰 {data['total']}\n"
            f"🏠 Samarqand, Pastdarg'om, Charxin\n"
        )

        await update.message.reply_text(
            "📍 Olib ketish manzili:\nSamarqand, Pastdarg'om, Charxin\n A'loqa 📞 +998915388499  Adminlar o'zlari a'loqaga chiqishadi va manzilni yetgazishadi. "
        )

        await update.message.reply_text(
            "🏠 Bosh menyu",
            reply_markup=MAIN_MENU
        )

    # 🔥 ENG MUHIM — ADMIN GA HAR DOIM YUBORILADI
    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=text_admin,
        reply_markup=InlineKeyboardMarkup(admin_keyboard)
    )

    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=f"📦 Pickup tayyor\nID: {order_id}"
    )

    # ===== TOZALASH =====
    carts[user_id] = {}
    context.user_data.clear()

# Application'ni qurishda quyidagi tartibda qo'shing:



# ─────────────────────────────────────────────
# SHABLON GENERATSIYA (openpyxl)
# ─────────────────────────────────────────────
def generate_shablon_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import io as _io

    wb = Workbook()

    # ── Yashirin ro'yxatlar varag'i ──
    ref_ws = wb.active
    ref_ws.title = "Royxatlar"
    lists = {
        "A": ("Jins",       ["O'g'il", "Qiz"]),
        "B": ("Fabrika",    ["Vodiy", "Xitoy", "Turkiya", "8-mart fabrika"]),
        "C": ("Fasl",       ["Yozgi", "Qishki", "Bahor", "Kuz"]),
        "D": ("Kategoriya", [
            "2 talik kiyim", "3 talik kiyim", "futbolka",
            "shim", "qalin kiyim", "shortik",
            "oyoq kiyim", "bosh kiyim", "ichki kiyim"
        ]),
    }
    for col, (header, values) in lists.items():
        ref_ws[f"{col}1"] = header
        ref_ws[f"{col}1"].font = Font(bold=True)
        for i, v in enumerate(values, start=2):
            ref_ws[f"{col}{i}"] = v
    ref_ws.sheet_state = "hidden"

    # ── Asosiy varaq ──
    ws = wb.create_sheet("Mahsulotlar", 0)

    HEADER_BG = "1F4E79"
    REQD_BG   = "D6E4F0"
    OPT_BG    = "EBF5FB"
    EXAMPLE_BG= "FFF9C4"
    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = [
        ("photo (file_id)",     38, True,  "/get_id → rasmni yuboring → kodni shu yerga"),
        ("Nomi",                22, True,  "Masalan: Bolalar kostyumi"),
        ("Jins",                12, True,  "Royxatdan tanlang"),
        ("Fabrika",             18, True,  "Royxatdan tanlang"),
        ("Fasl",                16, True,  "Royxatdan tanlang. Bir nechta: Yozgi,Bahor"),
        ("Kategoriya",          20, True,  "Royxatdan tanlang"),
        ("Razmer (sm)",         14, True,  "Santimetrda: 44 yoki 86-92"),
        ("Narx (so'm)",        16, True,  "Faqat raqam: 85000"),
        ("Tannarx (so'm)",     16, False, "Foyda hisoblash uchun: 60000"),
        ("Soni",                10, True,  "Nechta mavjud: 4"),
        ("Ranglar (ixtiyoriy)", 28, False, "Oq:4,Qora:3,Ko'k:2"),
        ("Izoh",                30, False, "Qo'shimcha ma'lumot"),
    ]

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 34

    for ci, (title, width, required, note) in enumerate(COLS, start=1):
        col_l = get_column_letter(ci)

        c = ws.cell(row=1, column=ci, value=title)
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

        n = ws.cell(row=2, column=ci, value=note)
        n.font = Font(italic=True, color="555555", size=9, name="Arial")
        n.fill = PatternFill("solid", start_color="F0F0F0")
        n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n.border = border

        ws.column_dimensions[col_l].width = width

    examples = [
        "1", "Bolalar sport kostyumi",
        "O'g'il", "Xitoy", "Yozgi", "2 talik kiyim",
        "86-92", "85000", "60000", "4", "Oq:2,Ko'k:2", "Engil material",
    ]
    ws.cell(row=3, column=1).value = "NAMUNA — bu qatorni o'chiring"
    ws.cell(row=3, column=1).font = Font(bold=True, color="B8860B", name="Arial")
    for ci, val in enumerate(examples, start=1):
        c = ws.cell(row=3, column=ci)
        if ci > 1:
            c.value = val
        c.fill = PatternFill("solid", start_color=EXAMPLE_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
        c.font = Font(italic=True, color="666666", size=10, name="Arial")

    for row in range(4, 104):
        for ci, (_, _, required, _) in enumerate(COLS, start=1):
            c = ws.cell(row=row, column=ci)
            c.fill = PatternFill("solid", start_color=REQD_BG if required else OPT_BG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
            c.font = Font(name="Arial", size=10)

    # Dropdownlar
    dv_jins = DataValidation(type="list", formula1="=Royxatlar!$A$2:$A$3", allow_blank=True, showDropDown=False)
    dv_jins.sqref = "C4:C103"
    ws.add_data_validation(dv_jins)

    dv_fab = DataValidation(type="list", formula1="=Royxatlar!$B$2:$B$5", allow_blank=True, showDropDown=False)
    dv_fab.sqref = "D4:D103"
    ws.add_data_validation(dv_fab)

    dv_kat = DataValidation(type="list", formula1="=Royxatlar!$D$2:$D$10", allow_blank=True, showDropDown=False)
    dv_kat.sqref = "F4:F103"
    ws.add_data_validation(dv_kat)

    dv_num = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
    dv_num.sqref = "H4:J103"
    ws.add_data_validation(dv_num)

    ws.freeze_panes = "A4"

    # Yoriqnoma varag'i
    yw = wb.create_sheet("Yoriqnoma")
    yw.column_dimensions["A"].width = 5
    yw.column_dimensions["B"].width = 24
    yw.column_dimensions["C"].width = 55
    yw.merge_cells("A1:C1")
    h = yw.cell(row=1, column=1, value="Kiym bot — Excel shablon yo'riqnomasi")
    h.font = Font(bold=True, size=14, color="1F4E79", name="Arial")

    steps = [
        ("📸", "1. Rasm kodi olish",   "Botga /get_id → rasmlarni yuboring → file_id larni kopiyalang"),
        ("📋", "2. Jadvalni to'ldiring","Har bir mahsulot = 1 qator. NAMUNA qatorini o'chiring"),
        ("📁", "3. .xlsx saqlang",      "Fayl → Saqlash → xlsx formatda"),
        ("📤", "4. Botga yuboring",     "Excel faylni botga yuboring → avtomatik DB ga yoziladi"),
        ("🎨", "Ranglar ustuni",        "Ixtiyoriy. Format: Oq:4,Qora:3,Ko'k:2"),
        ("📏", "Razmer ustuni",         "44  yoki  86-92  yoki  3-4 yosh"),
        ("🌸", "Fasl bir nechta bo'lsa","Vergul bilan: Yozgi,Bahor"),
    ]
    for i, (icon, title, desc) in enumerate(steps, start=3):
        yw.cell(row=i, column=1, value=icon).font = Font(size=13)
        t = yw.cell(row=i, column=2, value=title)
        t.font = Font(bold=True, size=11, name="Arial", color="1F4E79")
        d = yw.cell(row=i, column=3, value=desc)
        d.font = Font(size=10, name="Arial")
        d.alignment = Alignment(wrap_text=True)
        yw.row_dimensions[i].height = 28

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def shablon_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    await update.message.reply_text("⏳ Shablon tayyorlanmoqda...")
    data = generate_shablon_bytes()
    await update.message.reply_document(
        document=data,
        filename="kiym_shablon.xlsx",
        caption=(
            "📋 Excel shablon\n\n"
            "1️⃣ /get_id → rasmlarni yuboring → file_id larni oling\n"
            "2️⃣ Shu faylni to'ldiring (dropdown lar bor)\n"
            "3️⃣ NAMUNA qatorini o'chiring\n"
            "4️⃣ Faylni menga yuboring → DB ga yozaman"
        )
    )

app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("get_id", get_id_command))
app.add_handler(CommandHandler("done", done_command))
app.add_handler(CommandHandler("shablon", shablon_command))
app.add_handler(MessageHandler(filters.Document.ALL, excel_import_handler))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))
app.add_handler(MessageHandler(filters.PHOTO, photo_handler))
app.add_handler(MessageHandler(filters.CONTACT, contact_handler))
app.add_handler(MessageHandler(filters.LOCATION, location_handler))
app.add_handler(CallbackQueryHandler(button_handler, pattern=".*"))
load_products_from_db()
app.run_polling()
