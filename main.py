import json
import time
import io
import openpyxl
from telegram import KeyboardButton
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import CallbackQueryHandler
import asyncio
import os
import psycopg2
from telegram import InputMediaPhoto

DATABASE_URL = os.getenv("DATABASE_URL")

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()   # 🔥 SHU MUHIM
cur.execute("""
CREATE TABLE IF NOT EXISTS shop_products (
    id SERIAL PRIMARY KEY,
    photo TEXT,
    gender TEXT,
    origin TEXT,
    season TEXT,
    category TEXT,
    name TEXT,
    size TEXT,
    price TEXT,
    count INTEGER,
    reserved INTEGER DEFAULT 0
)
""")

cur.execute("UPDATE shop_products SET reserved = 0")
conn.commit()

cur.execute("DELETE FROM shop_products WHERE photo IS NULL OR photo = ''")
conn.commit()

cur.execute("""
CREATE TABLE IF NOT EXISTS shop_photos (
    id SERIAL PRIMARY KEY,
    file_id TEXT NOT NULL,
    created_at FLOAT
)
""")
conn.commit()

cur.execute("""
ALTER TABLE shop_products ADD COLUMN IF NOT EXISTS cost INTEGER DEFAULT 0
""")
conn.commit()


cur.execute("""
CREATE TABLE IF NOT EXISTS shop_orders (
    id SERIAL PRIMARY KEY,
    user_id BIGINT,
    cart TEXT,
    location TEXT,
    phone TEXT,
    total INTEGER,
    status TEXT,
    time FLOAT
)
""")

conn.commit()



cur.execute("""
CREATE TABLE IF NOT EXISTS shop_users (
    user_id BIGINT PRIMARY KEY
)
""")
conn.commit()

cur.execute("""
ALTER TABLE shop_users ADD COLUMN IF NOT EXISTS created_at FLOAT
""")
conn.commit()

TOKEN = os.getenv("TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

# ═══════════════════════════════════════════
# ADMIN INLINE KLAVIATURALAR
# ═══════════════════════════════════════════

def admin_main_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📦 Mahsulotlar", callback_data="adm_products"),
            InlineKeyboardButton("🛒 Buyurtmalar", callback_data="adm_orders"),
        ],
        [
            InlineKeyboardButton("📊 Statistika",  callback_data="adm_stats"),
            InlineKeyboardButton("📢 Reklama",     callback_data="adm_broadcast"),
        ],
        [
            InlineKeyboardButton("📸 Rasm yuklash", callback_data="adm_photo"),
            InlineKeyboardButton("📋 Shablon",      callback_data="adm_shablon"),
        ],
    ])

def admin_products_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("➕ Qo'shish",     callback_data="adm_add"),
            InlineKeyboardButton("📥 Excel import", callback_data="adm_import"),
        ],
        [
            InlineKeyboardButton("📋 Ro'yxat",      callback_data="adm_list_0"),
            InlineKeyboardButton("🗑 Hammasini o'chir", callback_data="adm_clear"),
        ],
        [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_back_main")],
    ])

def admin_add_gender_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("👦 O'g'il", callback_data="adm_gender_ogil"),
            InlineKeyboardButton("👧 Qiz",     callback_data="adm_gender_qiz"),
        ],
        [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_products")],
    ])

def admin_add_origin_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🇺🇿 Vodiy",   callback_data="adm_origin_Vodiy"),
            InlineKeyboardButton("🇨🇳 Xitoy",   callback_data="adm_origin_Xitoy"),
        ],
        [
            InlineKeyboardButton("🇹🇷 Turkiya", callback_data="adm_origin_Turkiya"),
            InlineKeyboardButton("🏭 8-mart",   callback_data="adm_origin_8-mart fabrika"),
        ],
        [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_add")],
    ])

def admin_add_season_kb(selected=None):
    if selected is None:
        selected = []
    def mark(s): return f"✅ {s}" if s in selected else s
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton(mark("Yozgi"),  callback_data="adm_season_Yozgi"),
            InlineKeyboardButton(mark("Qishki"), callback_data="adm_season_Qishki"),
        ],
        [
            InlineKeyboardButton(mark("Bahor"),  callback_data="adm_season_Bahor"),
            InlineKeyboardButton(mark("Kuz"),    callback_data="adm_season_Kuz"),
        ],
        [InlineKeyboardButton("✅ Tayyor", callback_data="adm_season_done")],
        [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_add")],
    ])

def admin_add_category_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("👕 2 talik",  callback_data="adm_cat_2 talik kiyim"),
            InlineKeyboardButton("👕 3 talik",  callback_data="adm_cat_3 talik kiyim"),
        ],
        [
            InlineKeyboardButton("👕 Futbolka", callback_data="adm_cat_futbolka"),
            InlineKeyboardButton("👖 Shim",     callback_data="adm_cat_shim"),
        ],
        [
            InlineKeyboardButton("🧥 Qalin",    callback_data="adm_cat_qalin kiyim"),
            InlineKeyboardButton("🩳 Shortik",  callback_data="adm_cat_shortik"),
        ],
        [
            InlineKeyboardButton("👟 Oyoq",     callback_data="adm_cat_oyoq kiyim"),
            InlineKeyboardButton("🧢 Bosh",     callback_data="adm_cat_bosh kiyim"),
        ],
        [
            InlineKeyboardButton("🩲 Ichki",    callback_data="adm_cat_ichki kiyim"),
        ],
        [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_add")],
    ])

def admin_clear_confirm_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Ha, o'chir", callback_data="clear_yes"),
            InlineKeyboardButton("❌ Yo'q",       callback_data="adm_products"),
        ],
    ])

def admin_product_item_kb(product_id, page=0, has_prev=False, has_next=False):
    nav = []
    if has_prev:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"adm_list_{page-1}"))
    if has_next:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"adm_list_{page+1}"))
    rows = [
        [
            InlineKeyboardButton("✏️ Tahrir", callback_data=f"edit_{product_id}"),
            InlineKeyboardButton("🗑 O'chir", callback_data=f"delete_{product_id}"),
        ],
    ]
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton("🔙 Orqaga", callback_data="adm_products")])
    return InlineKeyboardMarkup(rows)



product_locks = {}
products = []
carts = {}
order_id_counter = 1
ORIGINS = [
    "🇺🇿 Vodiy",
    "🇨🇳 Xitoy",
    "🇹🇷 Turkiya",
    "🏭 8-mart fabrika"
]
def get_filter_menu(user_data):
    g = user_data.get("filter_gender", "-")
    o = user_data.get("filter_origin", "-")
    s = user_data.get("filter_season", "-")

    return InlineKeyboardMarkup([

        # 🔥 JINS LABEL
        [InlineKeyboardButton(f"Jins: {g}", callback_data="empty")],

        [
            InlineKeyboardButton(f"👦 O‘g‘il {'✅' if g=='o‘g‘il' else ''}", callback_data="g_o‘g‘il"),
            InlineKeyboardButton(f"👧 Qiz {'✅' if g=='qiz' else ''}", callback_data="g_qiz"),
        ],

        # 🔥 FABRIKA LABEL
        [InlineKeyboardButton(f"Fabrika: {o}", callback_data="empty")],

        [
            InlineKeyboardButton(f"🇺🇿 Vodiy {'✅' if o=='Vodiy' else ''}", callback_data="o_Vodiy"),
            InlineKeyboardButton(f"🇨🇳 Xitoy {'✅' if o=='Xitoy' else ''}", callback_data="o_Xitoy"),
        ],
        [
            InlineKeyboardButton(f"🇹🇷 Turkiya {'✅' if o=='Turkiya' else ''}", callback_data="o_Turkiya"),
            InlineKeyboardButton(f"🏭 8-mart {'✅' if o=='8-mart fabrika' else ''}", callback_data="o_8-mart fabrika"),
        ],

        # 🔥 FASL LABEL
        [InlineKeyboardButton(f"Fasl: {s}", callback_data="empty")],

        [
            InlineKeyboardButton(f"☀️ Yozgi {'✅' if s=='Yozgi' else ''}", callback_data="s_Yozgi"),
            InlineKeyboardButton(f"❄️ Qishki {'✅' if s=='Qishki' else ''}", callback_data="s_Qishki"),
        ],
        [
            InlineKeyboardButton(f"🌸 Bahor {'✅' if s=='Bahor' else ''}", callback_data="s_Bahor"),
            InlineKeyboardButton(f"🍂 Kuz {'✅' if s=='Kuz' else ''}", callback_data="s_Kuz"),
        ],

        [
            InlineKeyboardButton("✅ Tanlash", callback_data="apply"),
            InlineKeyboardButton("🔄 Tozalash", callback_data="reset"),
        ]
    ])
def get_category_buttons(context):
    return [
        [f"👕 2 talik ({count_products(context,lambda p: p['category'].lower()=='2 talik kiyim')})",
         f"👕 3 talik ({count_products(context,lambda p: p['category'].lower()=='3 talik kiyim')})",
         f"👕 Futbolka ({count_products(context,lambda p: p['category'].lower()=='futbolka')})"],

         [f"👖 Shim ({count_products(context,lambda p: p['category'].lower()=='shim')})",
          f"🧥 Qalin ({count_products(context,lambda p: p['category'].lower()=='qalin kiyim')})",
         f"🩳 Shortik ({count_products(context,lambda p: p['category'].lower()=='shortik')})"],

        [f"👟 Oyoq ({count_products(context,lambda p: p['category'].lower()=='oyoq kiyim')})",
         f"🧢 Bosh ({count_products(context,lambda p: p['category'].lower()=='bosh kiyim')})",
         f"🩲 Ichki ({count_products(context,lambda p: p['category'].lower()=='ichki kiyim')})"],

        ["🔙 Orqaga", "🏠 Bosh menyu"]
    ]

def filter_check(p, context):

    # 🔥 gender
    g = context.user_data.get("filter_gender")
    if g:
        if g.lower() not in str(p.get("gender", "")).lower():
            return False

    # 🔥 origin
    o = context.user_data.get("filter_origin")
    if o:
        if o.lower() not in str(p.get("origin", "")).lower():
            return False

    # 🔥 category (YENGIL VA ISHONCHLI)
    c = context.user_data.get("filter_category")
    if c:
        user_cat = str(c).strip().lower()
        prod_cat = str(p.get("category", "")).strip().lower()

        if not prod_cat:
            return False

        # 🔥 faqat o‘xshashlikni tekshiramiz (soft)
        if user_cat not in prod_cat:
            return False

    # 🔥 season (TOZA)
    s = context.user_data.get("filter_season")
    if s:
        season = str(s).strip().lower()

        p_seasons = p.get("season", [])

        # har doim listga aylantiramiz
        if not isinstance(p_seasons, list):
            p_seasons = str(p_seasons).split(",")

        p_seasons = [str(x).strip().lower() for x in p_seasons]

        if season not in p_seasons:
            return False

    # 🔥 size
    if context.user_data.get("filter_size"):
        size_text = context.user_data.get("filter_size")
        if str(size_text).isdigit():
            size = int(size_text)
            raw = str(p.get("size") or "").lower().replace("sm", "").strip()

            if raw == "":
                print(f"DEBUG: {p['name']} - o'lcham bo'sh")
                return False

            if "-" in raw:
                parts = raw.split("-")
                if len(parts) >= 2 and parts[0].strip().isdigit() and parts[1].strip().isdigit():
                    s1, s2 = int(parts[0]), int(parts[1])
                    if not (s1 <= size <= s2):
                        print(f"DEBUG: {p['name']} - o'lcham diapazonga tushmadi ({s1}-{s2})")
                        return False
                else: return False
            else:
                if not raw.isdigit(): return False
                p_size = int(raw)
                if abs(p_size - size) > 1:
                    print(f"DEBUG: {p['name']} - o'lcham mos kelmadi (P:{p_size}, U:{size})")
                    return False

    # 🔥 mavjudlik (Eng ko'p xato shu yerda bo'ladi)
    available = p.get("count", 0) - p.get("reserved", 0)

    if available <= 0:
        return False

    # Agar barcha shartlardan o'tsa
    return True 
def clean_cart(user_id, context=None):
    
    now = time.time()

    # 🔥 BUYURTMA BOSILGAN BO‘LSA — TO‘XTAT
    if context and context.user_data.get("order_started"):
        return carts.get(user_id, {})

    cart = carts.get(user_id, {})
    new_cart = {}

    for pid, item in cart.items():
        if now - item["time"] < 7200:
            new_cart[pid] = item
        else:
            p = next((x for x in products if x["id"] == int(pid)), None)
            if p:
                p["reserved"] = max(0, p.get("reserved", 0) - item["qty"])

    carts[user_id] = new_cart
    return new_cart

def load_products_from_db():
    global products

    cur.execute("SELECT * FROM shop_products")
    rows = cur.fetchall()

    products = []

    for r in rows:
        products.append({
            "id": r[0],
            "photo": r[1],
            "gender": r[2],
            "origin": r[3],
            "season": r[4].split(",") if r[4] else [],
            "category": r[5],
            "name": r[6],
            "size": r[7],
            "price": r[8],
            "count": r[9],
            "reserved": r[10]
        })
def count_products(context, filter_func=None):
    total = 0
    for p in products:
        available = p["count"] - p.get("reserved", 0)

        if available > 0 and filter_check(p, context):
            if not filter_func or filter_func(p):
                total += available

    return total


ADMIN_MENU = ReplyKeyboardMarkup(
    [
        ["📊 Statistika", "📦 Buyurtmalar"],
        [ "📢 Reklama"],
        ["🏠 Bosh menyu"]
    ],
    resize_keyboard=True
)

MAIN_MENU = ReplyKeyboardMarkup(
    [
        ["🔍 Qidirish"],
        ["🛍 Kiyimlar", "🧺 Savat"],
        ["ℹ️ Yordam"]
    ],
    resize_keyboard=True
)

BACK_BUTTON = ["🔙 Orqaga"]
HOME_BUTTON = ["🏠 Bosh menyu"]
CART_BUTTON = ["🧺 Savat"]

CATEGORIES = [
    "👕 2 talik kiyim (dvoyka)",
    "👕 3 talik kiyim (troyka)",
    "👕 Futbolka(qizlarga ko‘ylak ham)",
    "👖 Shim",
    "🧥 Qalin kiyim",
    "🩳 Shortik(qizlarga yubka)",
    "👟 oyoq kiyim",
    "🧢 Bosh kiyim",
    "🩲 Ichki kiyim"
]

async def show_products(update, context, products, ADMIN_ID):

    page = context.user_data.get("page", 0)
    start = page * 4
    end = start + 4

    chunk = products[start:end]

    if not chunk:
        await update.message.reply_text("❌ Hech narsa topilmadi")
        return

    # 🔥 MEDIA VA VALID MAHSULOTLAR YIG'ISH
    media = []
    valid_products = []

    for i, p in enumerate(chunk):
        if not p.get("photo"):
            continue  # rasm yo'q bo'lsa skip

        media.append(
            InputMediaPhoto(
                media=p.get("photo"),
                caption=f"{i+1}) {p.get('size')}"
            )
        )
        valid_products.append(p)

    # 🔥 ALBUM YUBORISH
    if media:
        await update.message.reply_media_group(media)
    else:
        await update.message.reply_text("❌ Rasm topilmadi")
        return

    # 🔽 INFO + TUGMALAR
    for i, p in enumerate(valid_products):
        keyboard = [
            [InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p.get('id')}")]
        ]

        if update.effective_user.id == ADMIN_ID:
            keyboard.append([
                InlineKeyboardButton("✏️ Edit", callback_data=f"edit_{p.get('id')}"),
                InlineKeyboardButton("🗑 O‘chirish", callback_data=f"delete_{p.get('id')}")
            ])

        await update.message.reply_text(
            f"{i+1}) {p.get('name')}\n📏 {p.get('size')} sm\n💰 {p.get('price')}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # 🔽 NEXT / PREV
    nav = []

    if start > 0:
        nav.append(InlineKeyboardButton("⬅️ Orqaga", callback_data="prev"))

    if end < len(products):
        nav.append(InlineKeyboardButton("➡️ Keyingi", callback_data="next"))

    if nav:
        await update.message.reply_text(
            "📄 Sahifa",
            reply_markup=InlineKeyboardMarkup([nav])
        )

async def start_photo_flow(context, chat_id, file_id):
    # 🔥 bu photo_handler ichidagi boshlanish logikasi
    context.user_data.clear()
    context.user_data["photo"] = file_id
    context.user_data["step"] = "gender"

    keyboard = [["👦 O‘g‘il", "👧 Qiz"]]
    await context.bot.send_message(
        chat_id=chat_id,
        text="Kim uchun?",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
# START
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    context.user_data.clear()

    user_id = update.effective_user.id

    cur.execute("""
    INSERT INTO shop_users (user_id, created_at)
    VALUES (%s, %s)
    ON CONFLICT (user_id) DO NOTHING
    """, (user_id, time.time()))
    conn.commit()

    load_products_from_db()

    # 🔥 BIR MARTA JAVOB BERAMIZ
    if user_id == ADMIN_ID:
        await update.message.reply_text(
            "👑 Admin panel\nAssalomu aleykum AZIZJON AHMADOVICH",
            reply_markup=admin_main_kb()
        )
    else:
        await update.message.reply_text(
            "Assalomu alaykum 👋",
            reply_markup=MAIN_MENU
        )
# RASM QABUL (ADMIN)
async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if user_id != ADMIN_ID and not update.message.from_user.is_bot:
        return

    caption = (update.message.caption or "").strip()

    # ── SPLIT rejimi: caption = "split:4:2" ──
    if caption.lower().startswith("split:"):
        parts = caption[6:].strip().split(":")
        try:
            rows = int(parts[0]) if len(parts) > 0 else 3
            cols = int(parts[1]) if len(parts) > 1 else 3
        except:
            await update.message.reply_text("❌ Format: split:qator:ustun\nMasalan: split:4:2")
            return

        total = rows * cols
        await update.message.reply_text(
            f"⏳ Rasm {rows}×{cols}={total} bo'lakka bo'linmoqda..."
        )

        import io
        from PIL import Image
        from telegram import InputFile

        # Rasmni yuklab olish
        file = await update.message.photo[-1].get_file()
        buf = io.BytesIO()
        await file.download_to_memory(buf)
        buf.seek(0)
        img = Image.open(buf)
        W, H = img.size

        cell_w = W // cols
        cell_h = H // rows

        saved = 0
        results = []

        for row in range(rows):
            for col in range(cols):
                n = row * cols + col + 1

                # Rasmni kesish
                x1 = col * cell_w
                y1 = row * cell_h
                x2 = x1 + cell_w
                y2 = y1 + cell_h
                piece = img.crop((x1, y1, x2, y2))

                # Telegram ga yuborish
                piece_buf = io.BytesIO()
                piece.save(piece_buf, format="JPEG", quality=95)
                piece_buf.seek(0)

                sent = await update.message.reply_photo(
                    photo=piece_buf,
                    caption=f"#{n}"
                )
                file_id = sent.photo[-1].file_id

                # DB ga saqlash
                cur.execute(
                    "INSERT INTO shop_photos (file_id, created_at) VALUES (%s, %s) RETURNING id",
                    (file_id, time.time())
                )
                photo_num = cur.fetchone()[0]
                conn.commit()
                results.append(f"#{n} → raqam: <code>{photo_num}</code>")
                saved += 1

        result_text = "\n".join(results)
        await update.message.reply_text(
            f"✅ {saved} ta bo'lak saqlandi!\n\n{result_text}\n\nShu raqamlarni Excel ga yozing.",
            parse_mode="HTML"
        )
        return

    # ── /get_id rejimi — oddiy rasm, raqam qaytaradi ──
    if context.user_data.get("get_id_mode"):
        file_id = update.message.photo[-1].file_id
        cur.execute(
            "INSERT INTO shop_photos (file_id, created_at) VALUES (%s, %s) RETURNING id",
            (file_id, time.time())
        )
        photo_num = cur.fetchone()[0]
        conn.commit()
        await update.message.reply_text(
            f"✅ Rasm saqlandi!\n\n"
            f"📋 Raqam: <b>#{photo_num}</b>\n\n"
            f"Excel jadvalga <code>{photo_num}</code> raqamini yozing",
            parse_mode="HTML"
        )
        return

    # ── Oddiy mahsulot qo'shish rejimi ──
    context.user_data.clear()
    context.user_data["photo"] = update.message.photo[-1].file_id
    context.user_data["step"] = "gender"

    keyboard = [["👦 O'g'il", "👧 Qiz"]]
    await update.message.reply_text(
        "Kim uchun?",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )


# /get_id buyrug'i
async def get_id_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    context.user_data["get_id_mode"] = True
    await update.message.reply_text(
        "📸 Endi rasmlarni ketma-ket yuboring\n"
        "Har bir rasm uchun raqam qaytaraman (#1, #2, #3...)\n"
        "Shu raqamni Excel ga yozing.\n\n"
        "Tugagach /done yuboring.",
        reply_markup=ReplyKeyboardMarkup([["🏠 Bosh menyu"]], resize_keyboard=True)
    )


# /done buyrug'i
async def done_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("get_id_mode", None)
    await update.message.reply_text(
        "✅ file_id rejimi tugadi.",
        reply_markup=ADMIN_MENU if update.effective_user.id == ADMIN_ID else MAIN_MENU
    )


# Excel import validatsiya konstantlari
FASL_OPTIONS = {"yozgi", "qishki", "bahor", "kuz"}
VALID_CATEGORIES = {
    "2 talik kiyim", "3 talik kiyim", "futbolka", "shim",
    "qalin kiyim", "shortik", "oyoq kiyim", "bosh kiyim", "ichki kiyim"
}


async def excel_import_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return

    doc = update.message.document
    if not doc or not doc.file_name.endswith(".xlsx"):
        return

    await update.message.reply_text("⏳ Excel o'qilmoqda...")

    file = await context.bot.get_file(doc.file_id)
    file_bytes = await file.download_as_bytearray()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb["Mahsulotlar"]
    except Exception as e:
        await update.message.reply_text(f"❌ Fayl o'qilmadi: {e}")
        return

    added = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row[0] and not row[1]:
            continue
        if row[0] and "NAMUNA" in str(row[0]).upper():
            continue

        photo_raw = str(row[0]).strip() if row[0] else ""
        name      = str(row[1]).strip() if row[1] else ""

        # Raqam bo'lsa — photos jadvaldan file_id olish
        if photo_raw.isdigit():
            cur.execute("SELECT file_id FROM shop_photos WHERE id = %s", (int(photo_raw),))
            ph_row = cur.fetchone()
            if not ph_row:
                errors.append(f"Qator {row_num}: #{photo_raw} raqamli rasm topilmadi")
                continue
            photo = ph_row[0]
        else:
            photo = photo_raw  # to'g'ridan file_id yozilgan bo'lsa ham ishlaydi
        gender    = str(row[2]).strip() if row[2] else ""
        origin    = str(row[3]).strip() if row[3] else ""
        season    = str(row[4]).strip() if row[4] else ""
        category  = str(row[5]).strip().lower() if row[5] else ""
        size      = str(row[6]).strip() if row[6] else ""
        price_raw = str(row[7]).strip() if row[7] else ""
        cost_raw  = str(row[8]).strip() if row[8] else "0"
        count_raw = str(row[9]).strip() if row[9] else ""

        missing = []
        if not photo_raw: missing.append("photo raqami")
        if not name:      missing.append("nom")
        if not gender:    missing.append("jins")
        if not origin:    missing.append("fabrika")
        if not season:    missing.append("fasl")
        if not category:  missing.append("kategoriya")
        if not size:      missing.append("razmer")
        if not price_raw: missing.append("narx")
        if not count_raw: missing.append("soni")

        if missing:
            errors.append(f"Qator {row_num}: {', '.join(missing)} yetishmaydi")
            continue

        try:
            price_int = int("".join(filter(str.isdigit, price_raw)))
            price_str = f"{price_int:,}".replace(",", " ") + " so'm"
        except:
            errors.append(f"Qator {row_num}: narx noto'g'ri ({price_raw})")
            continue

        try:
            cost_int = int("".join(filter(str.isdigit, cost_raw))) if cost_raw and cost_raw != "0" else 0
        except:
            cost_int = 0

        try:
            count_int = int("".join(filter(str.isdigit, count_raw)))
        except:
            errors.append(f"Qator {row_num}: soni noto'g'ri ({count_raw})")
            continue

        season_parts = [s.strip().capitalize() for s in season.replace(",", " ").split()]
        valid_seasons = [s for s in season_parts if s.lower() in FASL_OPTIONS]
        if not valid_seasons:
            errors.append(f"Qator {row_num}: fasl noto'g'ri ({season})")
            continue
        season_db = ",".join(valid_seasons)

        if category not in VALID_CATEGORIES:
            errors.append(f"Qator {row_num}: kategoriya noto'g'ri ({category})")
            continue

        try:
            # Razmerlar vergul bilan bo'lsa — bir xil razmerlar guruhlash
            razmers = [r.strip() for r in size.split(",") if r.strip()]
            if not razmers:
                razmers = [size]

            # Bir xil razmerlarni guruhlash: {razmer: soni}
            from collections import Counter
            razmer_counts = Counter(razmers)

            for razmer, cnt in razmer_counts.items():
                cur.execute("""
                    INSERT INTO shop_products
                        (photo, gender, origin, season, category, name, size, price, count, reserved, cost)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s)
                """, (
                    photo, gender, origin, season_db,
                    category, name, razmer, price_str,
                    cnt, cost_int
                ))
                added += 1
        except Exception as e:
            errors.append(f"Qator {row_num}: DB xato — {e}")
            conn.rollback()
            continue

    conn.commit()
    load_products_from_db()

    msg = f"✅ {added} ta mahsulot qo'shildi!"
    if errors:
        err_text = "\n".join(errors[:10])
        if len(errors) > 10:
            err_text += f"\n... va yana {len(errors)-10} ta xato"
        msg += f"\n\n⚠️ Xatolar:\n{err_text}"

    await update.message.reply_text(msg)


async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not update.message:
            return

        text = update.message.text

        # ===== ADMIN INLINE STEPS =====
        adm_step = context.user_data.get("adm_step")

        if adm_step == "broadcast":
            cur.execute("SELECT user_id FROM shop_users")
            users = cur.fetchall()
            count = 0
            for u in users:
                try:
                    await context.bot.send_message(chat_id=u[0], text=text)
                    count += 1
                except:
                    pass
            context.user_data.pop("adm_step", None)
            await update.message.reply_text(
                f"✅ {count} ta foydalanuvchiga yuborildi",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔙 Admin panel", callback_data="adm_back_main")]
                ])
            )
            return

        if adm_step == "excel_import":
            # fayl kutilmoqda — hozir document handler ishlaydi
            await update.message.reply_text("📥 .xlsx faylni yuboring")
            return

        if adm_step == "add_photo_num":
            photo_num = text.strip()
            if photo_num.isdigit():
                cur.execute("SELECT file_id FROM shop_photos WHERE id = %s", (int(photo_num),))
                ph_row = cur.fetchone()
                if not ph_row:
                    await update.message.reply_text(f"❌ #{photo_num} raqamli rasm topilmadi\n📸 Avval rasmni yuklang")
                    return
                context.user_data["new_product"]["photo"] = ph_row[0]
            else:
                # file_id to'g'ridan yozilgan
                context.user_data["new_product"]["photo"] = photo_num
            context.user_data["adm_step"] = "add_name"
            p = context.user_data["new_product"]
            await update.message.reply_text(
                f"✅ Rasm qo'shildi\n\n📛 Mahsulot nomini yozing:\n(masalan: Bolalar sport kostyumi)"
            )
            return

        if adm_step == "add_name":
            context.user_data["new_product"]["name"] = text
            context.user_data["adm_step"] = "add_size"
            await update.message.reply_text("📏 Uzunligini yozing (sm):\n(masalan: 86-92 yoki 44)")
            return

        if adm_step == "add_size":
            context.user_data["new_product"]["size"] = text
            context.user_data["adm_step"] = "add_price"
            await update.message.reply_text("💰 Narxini yozing (so'mda):\n(masalan: 85000)")
            return

        if adm_step == "add_price":
            price_clean = "".join(filter(str.isdigit, text))
            if not price_clean:
                await update.message.reply_text("❌ Faqat raqam yozing")
                return
            price_int = int(price_clean)
            price_str = f"{price_int:,}".replace(",", " ") + " so'm"
            context.user_data["new_product"]["price"] = price_str
            context.user_data["adm_step"] = "add_cost"
            await update.message.reply_text("💸 Tannarxini yozing:\n(masalan: 60000)")
            return

        if adm_step == "add_cost":
            cost_clean = "".join(filter(str.isdigit, text))
            context.user_data["new_product"]["cost"] = int(cost_clean) if cost_clean else 0
            context.user_data["adm_step"] = "add_count"
            await update.message.reply_text("📦 Nechta bor?\n(masalan: 4)")
            return

        if adm_step == "add_count":
            if not text.strip().isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing")
                return
            p = context.user_data["new_product"]
            p["count"] = int(text.strip())
            # Ko'rsatish va tasdiqlash
            summary = (
                f"📋 MAHSULOT MA'LUMOTLARI:\n\n"
                f"👤 Jins: {p.get('gender','—')}\n"
                f"🏭 Fabrika: {p.get('origin','—')}\n"
                f"🌸 Fasl: {', '.join(p.get('seasons',[]))}\n"
                f"📂 Kategoriya: {p.get('category','—')}\n"
                f"📛 Nomi: {p.get('name','—')}\n"
                f"📏 O'lcham: {p.get('size','—')} sm\n"
                f"💰 Narx: {p.get('price','—')}\n"
                f"💸 Tannarx: {p.get('cost',0):,} so'm\n"
                f"📦 Soni: {p.get('count',0)}"
            )
            await update.message.reply_text(
                summary,
                reply_markup=InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("✅ Saqlash", callback_data="adm_save"),
                        InlineKeyboardButton("❌ Bekor",   callback_data="adm_products"),
                    ]
                ])
            )
            return

        # ===== ADMIN FLOW =====
        if context.user_data.get("step") == "gender":
            gender = text.replace("👦 ", "").replace("👧 ", "")
            context.user_data["gender"] = gender

            # 🔥 YANGI STEP
            context.user_data["step"] = "origin"

            keyboard = [
                ["🇺🇿 Vodiy", "🇨🇳 Xitoy"],
                ["🇹🇷 Turkiya", "🏭 8-mart fabrika"]
            ]

            await update.message.reply_text(
                "Qaysi ishlab chiqarish:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        


        elif context.user_data.get("step") == "broadcast":
            msg = text

            cur.execute("SELECT user_id FROM shop_users")
            users = cur.fetchall()

            count = 0

            for u in users:
                try:
                    await context.bot.send_message(chat_id=u[0], text=msg)
                    count += 1
                except:
                    pass

            await update.message.reply_text(f"✅ {count} ta userga yuborildi")

            context.user_data.clear()

        elif text == "🔍 Qidirish":
            context.user_data.clear()

            await update.message.reply_text(
                "🔎 Tanlang:\n\nJins: -\nFabrika: -\nFasl: -\n\n",
                reply_markup=get_filter_menu(context.user_data)
            )

            context.user_data["step"] = "inline_all"

        elif text == "📢 Reklama":
            if update.effective_user.id != ADMIN_ID:
                return

            context.user_data["step"] = "broadcast"
            await update.message.reply_text("📢 Yuboriladigan matnni yozing:")

        elif text == "📦 Buyurtmalar":
            if update.effective_user.id != ADMIN_ID:
                return

            cur.execute("SELECT id, total, status FROM shop_orders ORDER BY id DESC LIMIT 10")
            rows = cur.fetchall()

            if not rows:
                await update.message.reply_text("❌ Buyurtmalar yo‘q")
                return

            msg = "📦 So‘nggi buyurtmalar:\n\n"

            for r in rows:
                msg += f"🆔 {r[0]} | 💰 {r[1]} | 📌 {r[2]}\n"

            await update.message.reply_text(msg)

        elif text == "📊 Statistika":
            if update.effective_user.id != ADMIN_ID:
                return
            now = time.time()

            day = now - 86400
            week = now - 604800
            month = now - 2592000
            year = now - 31536000

            # 🔥 userlar
            cur.execute("SELECT COUNT(*) FROM shop_users")
            total_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM shop_users WHERE created_at >= %s", (day,))
            day_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM shop_users WHERE created_at >= %s", (week,))
            week_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM shop_users WHERE created_at >= %s", (month,))
            month_users = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM shop_users WHERE created_at >= %s", (year,))
            year_users = cur.fetchone()[0]

            # 🔥 jami buyurtma
            cur.execute("SELECT COUNT(*) FROM shop_orders")
            total_orders = cur.fetchone()[0]

            # 🔥 jami pul
            cur.execute("SELECT SUM(total) FROM shop_orders")
            total_money = cur.fetchone()[0] or 0

            cur.execute("SELECT cart FROM shop_orders")
            orders_data = cur.fetchall()

            total_profit = 0

            for row in orders_data:
                cart = json.loads(row[0])

                for pid, item in cart.items():
                    qty = item["qty"]

                    p = next((x for x in products if x["id"] == int(pid)), None)
                    if not p:
                        continue

                    price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                    cost = p.get("cost", 0)

                    # 🔥 SHU YERNI QO‘SH (ENG MUHIM)
                    if cost == 0:
                        continue

                    profit = (price - cost) * qty
                    total_profit += profit

            # 🔥 bugungi buyurtma
            cur.execute("""
            SELECT COUNT(*) FROM shop_orders 
            WHERE DATE(to_timestamp(time)) = CURRENT_DATE
            """)
            today_orders = cur.fetchone()[0]

            # 🔥 bugungi pul
            cur.execute("""
            SELECT SUM(total) FROM shop_orders 
            WHERE DATE(to_timestamp(time)) = CURRENT_DATE
            """)
            today_money = cur.fetchone()[0] or 0

            await update.message.reply_text(
                f"📊 STATISTIKA\n\n"

                f"👥 Jami user: {total_users}\n"
                f"📅 Bugun: {day_users}\n"
                f"📆 7 kun: {week_users}\n"
                f"🗓 1 oy: {month_users}\n"
                f"📈 1 yil: {year_users}\n\n"
                f"📈 FOYDA: {total_profit}\n\n"

                f"🧾 Jami buyurtma: {total_orders}\n"
                f"💰 Jami tushum: {total_money}\n\n"

                f"📦 Bugun buyurtma: {today_orders}\n"
                f"💵 Bugun tushum: {today_money}"
            )

        elif context.user_data.get("step") == "origin":
            origin = text.replace("🇺🇿 ", "").replace("🇨🇳 ", "").replace("🇹🇷 ", "").replace("🏭 ", "")
            context.user_data["origin"] = origin

            context.user_data["seasons"] = []
            context.user_data["step"] = "season"

            keyboard = [["☀️ Yozgi", "❄️ Qishki"], ["🌸 Bahor", "🍂 Kuz"]]

            await update.message.reply_text(
                "Fasl tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return    

        elif text == "✅ Tayyor" and context.user_data.get("step") == "season":
            context.user_data["step"] = "category"
        
            keyboard = get_category_buttons(context)
        
            await update.message.reply_text(
                "Kategoriya:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

        elif text == "ℹ️ Yordam":

            # 1-QISM
            await update.message.reply_text(
                "📏 1-QISM: Kiyimni qanday o‘lchash\n\n"
                "Bolaga mos eski kiyimni oling, stolga tekis qo‘ying va yuqoridan pastgacha uzunligini santimetrda o‘lchang. "
                "Masalan: 44 sm chiqdi. Shu o‘lcham eng muhim, chunki bot aynan shu bo‘yicha ishlaydi.\n\n"
                "Endi shu raqamni eslab qoling, chunki keyingi qadamda aynan shu orqali qidiruv qilasiz."
            )

            # 2-QISM
            await update.message.reply_text(
                "🔎 2-QISM: Qidirish va tanlash\n\n"
                "Botga 44 yozsangiz → 43, 44, 45 sm kiyimlar chiqadi. "
                "Bu sizga yaqin o‘lchamlarni ko‘rsatadi.\n\n"
                "Kattaroq kerak bo‘lsa → 46 yozing.\n"
                "Kichikroq kerak bo‘lsa → 42 yozing.\n\n"
                "Har bir kiyim ostida uzunligi yozilgan bo‘ladi, shu raqamga qarab tanlang."
            )

            # 3-QISM
            await update.message.reply_text(
                "🛒 3-QISM: Buyurtma berish\n\n"
                "Yoqgan kiyimni tanlab 🛒 Savatga qo‘shing.\n"
                "🧺 Savat ga kiring.\n"
                "🚚 Buyurtma berish ni bosing.\n"
                "Telefon raqamingizni yozing.\n\n"
                "Shu bilan buyurtma tugaydi va siz bilan bog‘lanishadi."
            )                   
        elif context.user_data.get("step") == "size_season" and text in ["☀️ Yozgi","❄️ Qishki","🌸 Bahor","🍂 Kuz"]:
            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "")
            context.user_data["filter_season"] = season
            context.user_data["step"] = "size_category"

        
            keyboard = get_category_buttons(context)
            await update.message.reply_text(
                "Kategoriya tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            ) 
            return  

        elif text == "❌ Lokatsiya ishlamayapti":
            user_id = update.effective_user.id
            cart = carts.get(user_id, {})

            if not cart:
                await update.message.reply_text("❌ Savat bo‘sh")
                return

            total = 0
            for pid, item in cart.items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                total += price * qty

            final = total + 0

            # 🔥 ENG MUHIM — TEMP ORDER
            context.user_data["temp_order"] = {
                "cart": cart,
                "location": {},   # 🔥 bo‘sh dict (None emas!)
                "total": final,
                "type": "delivery"
            }

            context.user_data["order_step"] = "phone"

            await update.message.reply_text(
        "📞 Telefon raqamingizni yozing:"
            )
        elif context.user_data.get("order_step") == "manual_location":
            address = text

            user_id = update.effective_user.id
            cart = carts.get(user_id, {})

            total = 0
            for pid, item in cart.items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                total += price * qty

            delivery = 0
            final = total + delivery

            context.user_data["temp_order"] = {
                "cart": cart,
                "location": {"text": address},
                "total": final,
                "type": "delivery"
            }

            context.user_data["order_step"] = "phone"

            await update.message.reply_text("📞 Telefon yuboring:")    

        elif text == "🔙 Orqaga":
            step = context.user_data.get("step")

        
            # 🔹 choose_type → gender
            if step == "choose_type":
                context.user_data["step"] = "user_gender"

                keyboard = [
                    ["👦 O‘g‘il", "👧 Qiz"],
                    ["🔙 Orqaga", "🏠 Bosh menyu"]
                ]

                await update.message.reply_text(
                    "Kim uchun:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )

            # 🔹 user_category → user_season
            elif step == "user_category":
                context.user_data["step"] = "user_season"

                keyboard = [
                    ["☀️ Yozgi","❄️ Qishki"],
                    ["🌸 Bahor","🍂 Kuz"],
                    ["🔙 Orqaga", "🏠 Bosh menyu"]
                ]

                await update.message.reply_text(
                    "Fasl tanlang:",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )

            # 🔹 user_season → choose_type
            elif step == "user_season":
                context.user_data["step"] = "choose_type"

                keyboard = [
                    ["📂 Umumiy"],
                    ["🔙 Orqaga", "🏠 Bosh menyu"]
                ]

                await update.message.reply_text(
                    "Qanday qidirasiz?",
                    reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                )

            # 🔹 default → bosh menyu
            else:
                context.user_data.clear()
                await update.message.reply_text("🏠 Bosh menyu", reply_markup=MAIN_MENU)

            return
        
        elif context.user_data.get("step") == "season":
            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "")
        
            if "seasons" not in context.user_data:
                context.user_data["seasons"] = []
        
            if season not in context.user_data["seasons"]:
                context.user_data["seasons"].append(season)
        
            keyboard = [
                ["☀️ Yozgi","❄️ Qishki"],
                ["🌸 Bahor","🍂 Kuz"],
                ["✅ Tayyor"]
            ]
        
            await update.message.reply_text(
                f"Tanlangan: {', '.join(context.user_data['seasons'])}",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
        
        elif context.user_data.get("step") == "category":

            category = text.split("(")[0]
            category = category.replace("👕","").replace("👖","").replace("🧥","") \
                            .replace("🩳","").replace("👟","").replace("🧢","").replace("🩲","").strip().lower()

            # 🔥 TO‘G‘RI NOMGA O‘TKAZAMIZ
            if "2 talik" in category:
                category = "2 talik kiyim"
            elif "3 talik" in category:
                category = "3 talik kiyim"
            elif "futbolka" in category:
                category = "futbolka"
            elif "shim" in category:
                category = "shim"
            elif "qalin" in category:
                category = "qalin kiyim"
            elif "shortik" in category:
                category = "shortik"
            elif "oyoq" in category:
                category = "oyoq kiyim"
            elif "bosh" in category:
                category = "bosh kiyim"
            elif "ichki" in category:
                category = "ichki kiyim"

            context.user_data["category"] = category.strip().lower()
            context.user_data["step"] = "name"

            await update.message.reply_text("Nomini yozing:")
            return

        elif context.user_data.get("step") == "all_season":

            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "").strip().lower()

            found = False

            for p in products:
                try:
                    p_seasons = p.get("season", [])

                    if not isinstance(p_seasons, list):
                        p_seasons = str(p_seasons).split(",")

                    p_seasons = [str(s).strip().lower() for s in p_seasons]

                    if season in p_seasons:

                        photo = p.get("photo")
                        if not photo:
                            continue

                        found = True

                        await update.message.reply_photo(
                            photo=photo,
                            caption=f"{p.get('name','')}\n📏 {p.get('size','')}\n💰 {p.get('price','')}"
                        )

                except Exception as e:
                    print("XATO:", p, e)
                    continue

            if not found:
                await update.message.reply_text("❌ Mahsulot yo‘q")

            context.user_data.clear()


        elif context.user_data.get("step") == "name":
            context.user_data["name"] = text
            context.user_data["step"] = "size"

            await update.message.reply_text("📏 Uzunlik yozing (sm) (masalan 40):")
            return
        elif context.user_data.get("step") == "size":
            size = text.strip()

            if not size.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing (masalan 44)")
                return

            context.user_data["size"] = size
            context.user_data["step"] = "price"

            await update.message.reply_text("Narxni yozing:")
                    
        elif context.user_data.get("step") == "price":
            price = text.replace(" ", "").replace("so'm","").replace("soʻm","")

            if not price.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing (masalan 50000)")
                return

            price = int(price)
            price = f"{price:,}".replace(",", " ")

            context.user_data["price"] = price + " so‘m"
            context.user_data["step"] = "cost"

            await update.message.reply_text("💸 Tannarxni yozing (masalan 30000):")

        elif context.user_data.get("step") == "cost":
            cost = text.replace(" ", "")

            if not cost.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing")
                return

            context.user_data["cost"] = int(cost)
            context.user_data["step"] = "count"

            await update.message.reply_text("📦 Nechta bor? (masalan 4):")

        elif context.user_data.get("step") == "count":
            if not text.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing")
                return

            context.user_data["count"] = int(text)

            # 🔥 xavfsiz olish
            photo = context.user_data.get("photo")
            gender = context.user_data.get("gender")
            origin = context.user_data.get("origin")
            seasons = context.user_data.get("seasons", [])
            category = context.user_data.get("category")
            name = context.user_data.get("name")
            size = context.user_data.get("size")
            price = context.user_data.get("price")
            cost = context.user_data.get("cost", 0)
            count = context.user_data.get("count")

            if context.user_data.get("mode") == "edit":
                pid = context.user_data.get("edit_product_id")

                cur.execute("""
                UPDATE shop_products
                SET photo=%s, gender=%s, origin=%s, season=%s,
                    category=%s, name=%s, size=%s, price=%s, count=%s
                WHERE id=%s
                """, (
                    photo, gender, origin,
                    ",".join(seasons),
                    category, name, size, price, count,
                    pid
                ))

                await update.message.reply_text("✅ Tahrirlandi!")
            else:
                cur.execute("""
                INSERT INTO shop_products (
                    photo, gender, origin, season, category, name, size, price, count, reserved, cost
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    photo,
                    gender,
                    origin,
                    ",".join(seasons),
                    category,
                    name,
                    size,
                    price,
                    count,
                    0,
                    context.user_data.get("cost", 0)
                ))

                await update.message.reply_text("✅ Qo‘shildi!")

            conn.commit()
            load_products_from_db()
            context.user_data.clear()
        elif context.user_data.get("step") == "size_filter":
            size = text.replace(" ", "")
            context.user_data["filter_size"] = size
            context.user_data["step"] = "size_season"

            keyboard = [
                ["☀️ Yozgi","❄️ Qishki"],
                ["🌸 Bahor","🍂 Kuz"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Fasl tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

            # ===== USER FLOW =====
        elif text == "🛍 Kiyimlar":
            context.user_data.clear() 
            keyboard = [["👦 O‘g‘il", "👧 Qiz"],["🔙 Orqaga", "🏠 Bosh menyu"]]
            await update.message.reply_text("Tanlang:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))

        # 👦 / 👧
        elif text in ["👦 O‘g‘il", "👧 Qiz"] and context.user_data.get("step") != "gender":
            gender = text.replace("👦 ", "").replace("👧 ", "")
            context.user_data["filter_gender"] = gender

            context.user_data["step"] = "origin_select"

            keyboard = [
                ["🇺🇿 Vodiy", "🇨🇳 Xitoy"],
                ["🇹🇷 Turkiya", "🏭 8-mart fabrika"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Qaysi ishlab chiqarish:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

        elif text == "📂 Umumiy":

            context.user_data["step"] = "user_season"

            keyboard = [
                ["☀️ Yozgi","❄️ Qishki"],
                ["🌸 Bahor","🍂 Kuz"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Fasl tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )


        elif text == "🧺 Savat":
            load_products_from_db()
            user_id = update.effective_user.id
            cart = clean_cart(user_id, context)

            
            now = time.time()
            new_cart = {}

            for pid, item in cart.items():
                if now - item["time"] < 7200:
                    new_cart[pid] = item
                else:
                    qty = item["qty"]
                    p = next((x for x in products if x["id"] == int(pid)), None)
                    if p:
                        p["reserved"] = max(0, p.get("reserved", 0) - qty)

            carts[user_id] = new_cart
            cart = new_cart

            if not cart:
                await update.message.reply_text("🧺 Savat bo‘sh")
                return

            msg = "🧺 Savat:\n\n"
            total = 0
            keyboard = []

            for pid, item in cart.items():
                qty = item["qty"]

                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

                summa = price * qty
                total += summa

                msg += f"{p['name']} x{qty} = {summa}\n"

                keyboard.append([
                    InlineKeyboardButton("❌", callback_data=f"del_{pid}")
                ])

            msg += f"\n💰 Jami: {total}"

            keyboard.append([InlineKeyboardButton("🚚 Buyurtma", callback_data="checkout")])
            keyboard.append([InlineKeyboardButton("🔙 Orqaga", callback_data="back")])

            await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard))

        # 🌦 FASL
        elif text in ["☀️ Yozgi","❄️ Qishki","🌸 Bahor","🍂 Kuz"]:
            season = text.replace("☀️ ", "").replace("❄️ ", "").replace("🌸 ", "").replace("🍂 ", "")
            context.user_data["filter_season"] = season
            context.user_data["step"] = "user_category"

            keyboard = get_category_buttons(context)
            await update.message.reply_text(
                "Kategoriya tanlang:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return
        
        elif context.user_data.get("step") == "origin_select":
            origin = text.replace("🇺🇿 ", "").replace("🇨🇳 ", "").replace("🇹🇷 ", "").replace("🏭 ", "")
            context.user_data["filter_origin"] = origin

            context.user_data["step"] = "choose_type"

            keyboard = [
                ["📂 Umumiy"],
                ["🔙 Orqaga", "🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "Qanday qidirasiz?",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )
            return


        elif "(" in text and context.user_data.get("step") == "user_category":

            category = text.split("(")[0]
            category = category.replace("👕","").replace("👖","").replace("🧥","") \
                            .replace("🩳","").replace("👟","").replace("🧢","").replace("🩲","").strip().lower()

            if "2 talik" in category:
                category = "2 talik kiyim"
            elif "3 talik" in category:
                category = "3 talik kiyim"
            elif "futbolka" in category:
                category = "futbolka"
            elif "shim" in category:
                category = "shim"
            elif "qalin" in category:
                category = "qalin kiyim"
            elif "shortik" in category:
                category = "shortik"
            elif "oyoq" in category:
                category = "oyoq kiyim"
            elif "bosh" in category:
                category = "bosh kiyim"
            elif "ichki" in category:
                category = "ichki kiyim"

            context.user_data["filter_category"] = category

            # 🔥 FILTER
            filtered = []

            for p in products:

                # kategoriya
                if category not in str(p.get("category", "")).lower():
                    continue

                # origin
                origin = context.user_data.get("filter_origin")
                if origin:
                    if origin.lower() not in str(p.get("origin", "")).lower():
                        continue

                # season
                season = context.user_data.get("filter_season")
                if season:
                    p_seasons = p.get("season", [])

                    if not isinstance(p_seasons, list):
                        p_seasons = str(p_seasons).split(",")

                    p_seasons = [s.strip().lower() for s in p_seasons]

                    if season.lower() not in p_seasons:
                        continue

                # mavjudlik
                available = p.get("count", 0) - p.get("reserved", 0)
                if available <= 0:
                    continue

                filtered.append(p)

            # ❗ BU YERDA ENDI FILTERNI TEKSHIRAMIZ
            if not filtered:
                await update.message.reply_text("❌ Mos mahsulot topilmadi")
                return

            context.user_data["filtered"] = filtered
            context.user_data["i"] = 0

            p = filtered[0]
            photo = p.get("photo")

            keyboard = [
                [
                    InlineKeyboardButton("⬅️", callback_data="prev_one"),
                    InlineKeyboardButton("➡️", callback_data="next_one")
                ],
                [InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p.get('id')}")]
            ]

            # 🔥 RASM BOR/YO‘Q HOLAT
            if photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=f"1/{len(filtered)}\n\n{p.get('name')}\n📏 {p.get('size')}\n💰 {p.get('price')}",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            else:
                await update.message.reply_text(
                    f"1/{len(filtered)}\n\n{p.get('name')}\n📏 {p.get('size')}\n💰 {p.get('price')}\n\n⚠️ Rasm yo‘q",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
        elif text == "🚚 Buyurtma berish":
            keyboard = [["🚚 Dastavka", "📍 Olib ketish"],["🔙 Orqaga", "🏠 Bosh menyu"]]

            await update.message.reply_text(
                "Qanday olasiz?",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )   
        elif text == "🚚 Dastavka":
            context.user_data["order_step"] = "location"
            context.user_data["order_type"] = "delivery"

            keyboard = [[KeyboardButton("📍 Lokatsiya yuborish", request_location=True)],
            ["❌ Lokatsiya ishlamayapti"],
            ["🏠 Bosh menyu"]
        ]

            await update.message.reply_text(
                " Dastavka narxi taxminan 20 000-50 000 so‘m atrofida bo‘ladi\n 📍 Lokatsiyangizni yuboring va \n⏳ Iltimos bir oz kuting... yoki lokatsiya ishlamasa pastdagi tugmani bosing:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )  

        elif context.user_data.get("order_step") == "phone":
            phone = text.strip().replace(" ", "")

            if phone.isdigit() and len(phone) == 9:
                phone = "+998" + phone
            elif phone.startswith("+998") and len(phone) == 13:
                pass
            else:
                await update.message.reply_text("❌ Noto‘g‘ri raqam!")
                return

            data = context.user_data.get("temp_order")
            if not data:
                await update.message.reply_text("❌ Xatolik")
                return

            user_id = update.effective_user.id
            cur.execute("""
            INSERT INTO shop_orders (user_id, cart, location, phone, total, status, time)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
            """, (
                user_id,
                json.dumps(data["cart"]),
                json.dumps(data["location"]),
                phone,
                data["total"],
                "new",
                time.time()
            ))

            order_id = str(cur.fetchone()[0])
            conn.commit()

            # ===== MAHSULOTNI KAMAYTIRISH =====
            for pid, item in data["cart"].items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if p:
                    p["count"] -= qty
                    p["reserved"] = max(0, p.get("reserved", 0) - qty)
                    # 🔥 DB ga ham yozamiz
                    cur.execute(
                        "UPDATE shop_products SET count = count - %s, reserved = GREATEST(0, reserved - %s) WHERE id = %s",
                        (qty, qty, p["id"])
                    )
            conn.commit()
            #save_products()

            # ===== USERGA MAHSULOT =====
            for pid, item in data["cart"].items():
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue
                qty = item["qty"]

                await context.bot.send_photo(
                    chat_id=user_id,
                    photo=p["photo"],
                    caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
                )

            # ===== USER STATUS =====
            if data.get("type") == "delivery":
                await update.message.reply_text(
                    "🚚 Buyurtma qabul qilindi",
                    reply_markup=MAIN_MENU
                )
            else:
                await update.message.reply_text(
                    "📍 Olib ketish manzili:\nSamarqand, Pastdarg‘om, Charxin\nA'loqa 📞 +998915388499  Adminlar o'zlari a'loqaga chiqishadi va manzilni yetgazishadi. " 
                    ,
                    reply_markup=MAIN_MENU
                )

                #await context.bot.send_location(
                #   chat_id=user_id,
                #  latitude=39.690149,
                # longitude=66.824828
                #)

                await update.message.reply_text(
                    "🏠 Bosh menyu",
                    reply_markup=MAIN_MENU
                )

            # ===== ADMIN TUGMALAR =====
            admin_keyboard = [
                [InlineKeyboardButton("📞 Aloqa", callback_data=f"contact_{order_id}")],
                [InlineKeyboardButton("🚚 Yetkazishni boshlash", callback_data=f"deliver_{order_id}")],
                [InlineKeyboardButton("✅ Yakunlandi", callback_data=f"done_{order_id}")],
                [InlineKeyboardButton("❌ Bekor", callback_data=f"cancel_{order_id}")]
            ]

            # ===== ADMINGA MAHSULOT =====
            for pid, item in data["cart"].items():
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue
                qty = item["qty"]

                await context.bot.send_photo(
                    chat_id=ADMIN_ID,
                    photo=p["photo"],
                    caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
                )

            # ===== ADMINGA UMUMIY INFO =====
            if data.get("type") == "delivery":
                text_admin = f"🚚 DASTAVKA\n📞 {phone}\n💰 {data['total']}"
            else:
                text_admin = f"📍 OLIB KETISH\n📞 {phone}\n💰 {data['total']}"

            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=text_admin,
                reply_markup=InlineKeyboardMarkup(admin_keyboard)
            )

            # ===== TOZALASH =====
            carts[user_id] = {}
            context.user_data.clear()

        elif text == "📍 Olib ketish":
            context.user_data["order_step"] = "phone"

            user_id = update.effective_user.id
            cart = carts.get(user_id, {})

            if not cart:
                await update.message.reply_text("❌ Savat bo‘sh")
                return

            total = 0
            for pid, item in cart.items():
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if not p:
                    continue

                price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))
                total += price * qty

            context.user_data["temp_order"] = {
                "cart": cart,
                "location": None,
                "total": total,
                "type": "pickup"
            }

            keyboard = [
                [KeyboardButton("📞 Telefon yuborish", request_contact=True)],
                ["🏠 Bosh menyu"]
            ]

            await update.message.reply_text(
                "📞 Telefon raqamingizni yuboring yoki yozing:",
                reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            )

        elif context.user_data.get("step") == "inline_category" and "(" in text:

            category = text.split("(")[0]
            category = category.replace("👕","").replace("👖","").replace("🧥","") \
                .replace("🩳","").replace("👟","").replace("🧢","").replace("🩲","").strip().lower()

            # 🔥 to‘g‘ri nomga o‘tkazamiz
            if "2 talik" in category:
                category = "2 talik kiyim"
            elif "3 talik" in category:
                category = "3 talik kiyim"
            elif "futbolka" in category:
                category = "futbolka"
            elif "shim" in category:
                category = "shim"
            elif "qalin" in category:
                category = "qalin kiyim"
            elif "shortik" in category:
                category = "shortik"
            elif "oyoq" in category:
                category = "oyoq kiyim"
            elif "bosh" in category:
                category = "bosh kiyim"
            elif "ichki" in category:
                category = "ichki kiyim"

            context.user_data["filter_category"] = category
            context.user_data["step"] = "write_size"

            await update.message.reply_text(
                "📏 Mahsulot uzunligini yozing (sm)\nMasalan: 44\n Bu kiyimdagi o‘lcham (razmer) emas maxsulotning uzunligi. Tushunmasangiz yordam bo‘limiga o‘ting bosh menyudan "
            )
        elif context.user_data.get("step") == "write_size":

            size = text.strip()

            if not size.isdigit():
                await update.message.reply_text("❌ Faqat raqam yozing (masalan 44)")
                return

            context.user_data["filter_size"] = size

            found = False

            for p in products:
                if filter_check(p, context):
                    found = True

                    keyboard = [
                        [InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p['id']}")]
                    ]

                    await update.message.reply_photo(
                        photo=p["photo"],
                        caption=f"{p['name']}\n📏 {p['size']} sm\n💰 {p['price']}",
                        reply_markup=InlineKeyboardMarkup(keyboard)
                    )

            # ❗ AGAR TOPILMASA
            if not found:
                await update.message.reply_text(
                    "❌ Mos mahsulot topilmadi.\n\nBoshqa razmer yozing (masalan 42, 46)"
                )
                return  # 🔥 MUHIM — step o‘chmaydi

            # 🔥 FAqat topilganda tozalaymiz
            context.user_data.clear()
        elif text == "🏠 Bosh menyu":
            context.user_data.clear()

            await update.message.reply_text(
                "🏠 Bosh menyu",
                reply_markup=MAIN_MENU
            )
            return

    except Exception as e:
        print("XATO:", e)
        await update.message.reply_text("❌ Xatolik yuz berdi")
        context.user_data.clear()

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    data = query.data



# ===== ADMIN CALLBACKS =====
    if data == "adm_back_main":
        await query.message.edit_text(
            "👑 Admin panel\nAssalomu aleykum AZIZJON AHMADOVICH",
            reply_markup=admin_main_kb()
        )
        return

    if data == "adm_products":
        await query.message.edit_text(
            "📦 Mahsulotlar boshqaruvi",
            reply_markup=admin_products_kb()
        )
        return

    if data == "adm_orders":
        cur.execute("SELECT id, phone, total, status FROM shop_orders ORDER BY id DESC LIMIT 10")
        rows = cur.fetchall()
        if not rows:
            await query.message.edit_text(
                "🛒 Hozircha buyurtmalar yo'q",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_back_main")]
                ])
            )
            return
        text = "🛒 So'nggi buyurtmalar:\n\n"
        kb = []
        for r in rows:
            text += f"🆔 {r[0]} | 📞 {r[1]} | 💰 {r[2]} | {r[3]}\n"
            kb.append([InlineKeyboardButton(
                f"#{r[0]} — {r[1]} ({r[2]} so'm)",
                callback_data=f"adm_order_{r[0]}"
            )])
        kb.append([InlineKeyboardButton("🔙 Orqaga", callback_data="adm_back_main")])
        await query.message.edit_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith("adm_order_"):
        order_id = data.split("_")[2]
        cur.execute("SELECT user_id, cart, phone, total, status FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()
        if not row:
            await query.answer("Topilmadi", show_alert=True)
            return
        uid, cart_json, phone, total, status = row
        cart = json.loads(cart_json)
        items_text = ""
        for pid, item in cart.items():
            p = next((x for x in products if x["id"] == int(pid)), None)
            if p:
                items_text += f"• {p['name']} x{item['qty']}\n"
        text = (
            f"🆔 Buyurtma #{order_id}\n"
            f"📞 {phone}\n"
            f"💰 {total} so'm\n"
            f"📌 {status}\n\n"
            f"{items_text}"
        )
        # status ga qarab tugmalar
        cur.execute("SELECT location FROM shop_orders WHERE id=%s", (order_id,))
        loc_row = cur.fetchone()
        loc = json.loads(loc_row[0]) if loc_row and loc_row[0] else {}
        order_type = "delivery" if loc and "lat" in loc else "pickup"
        if order_type == "delivery":
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("📞 Aloqa",     callback_data=f"contact_{order_id}"),
                    InlineKeyboardButton("🚚 Jo'natish", callback_data=f"send_{order_id}"),
                ],
                [
                    InlineKeyboardButton("✅ Yakunla",   callback_data=f"done_{order_id}"),
                    InlineKeyboardButton("❌ Bekor",     callback_data=f"cancel_{order_id}"),
                ],
                [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_orders")],
            ])
        else:
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("📞 Aloqa",      callback_data=f"contact_{order_id}"),
                    InlineKeyboardButton("📦 Tasdiqlash", callback_data=f"confirm_{order_id}"),
                ],
                [
                    InlineKeyboardButton("✅ Yakunla",    callback_data=f"done_{order_id}"),
                    InlineKeyboardButton("❌ Bekor",      callback_data=f"cancel_{order_id}"),
                ],
                [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_orders")],
            ])
        await query.message.edit_text(text, reply_markup=kb)
        return

    if data == "adm_stats":
        now = time.time()
        day = now - 86400
        week = now - 604800
        month = now - 2592000
        cur.execute("SELECT COUNT(*) FROM shop_users")
        total_users = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM shop_users WHERE created_at >= %s", (day,))
        day_users = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM shop_users WHERE created_at >= %s", (week,))
        week_users = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM shop_orders")
        total_orders = cur.fetchone()[0]
        cur.execute("SELECT SUM(total) FROM shop_orders")
        total_money = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*) FROM shop_orders WHERE DATE(to_timestamp(time)) = CURRENT_DATE")
        today_orders = cur.fetchone()[0]
        cur.execute("SELECT SUM(total) FROM shop_orders WHERE DATE(to_timestamp(time)) = CURRENT_DATE")
        today_money = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*) FROM shop_products")
        total_products = cur.fetchone()[0]
        await query.message.edit_text(
            f"📊 STATISTIKA\n\n"
            f"👥 Jami foydalanuvchi: {total_users}\n"
            f"📅 Bugun yangi: {day_users}\n"
            f"📆 7 kun: {week_users}\n\n"
            f"📦 Jami mahsulot: {total_products}\n\n"
            f"🧾 Jami buyurtma: {total_orders}\n"
            f"💰 Jami tushum: {total_money:,} so'm\n\n"
            f"📦 Bugun buyurtma: {today_orders}\n"
            f"💵 Bugun tushum: {today_money:,} so'm",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_back_main")]
            ])
        )
        return

    if data == "adm_broadcast":
        context.user_data["adm_step"] = "broadcast"
        await query.message.edit_text(
            "📢 Yubormoqchi bo'lgan xabarni yozing:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Bekor", callback_data="adm_back_main")]
            ])
        )
        return

    if data == "adm_photo":
        await query.message.edit_text(
            "📸 RASM YUKLASH\n\n"
            "1️⃣ Oddiy rasm → raqam olasiz\n"
            "2️⃣ Katta rasm qirqish uchun:\n"
            "   Caption ga yozing: <code>split:4:2</code>\n"
            "   (4 qator × 2 ustun = 8 bo'lak)\n\n"
            "Rasmni yuboring 👇",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_back_main")]
            ])
        )
        context.user_data["get_id_mode"] = True
        return

    if data == "adm_shablon":
        await query.answer("⏳ Tayorlanmoqda...")
        file_bytes = generate_shablon_bytes()
        await query.message.reply_document(
            document=file_bytes,
            filename="kiym_shablon.xlsx",
            caption=(
                "📋 Excel shablon\n\n"
                "1️⃣ /get_id yoki 📸 Rasm yuklash → rasmlarni yuboring\n"
                "2️⃣ Shablonni to'ldiring\n"
                "3️⃣ NAMUNA qatorini o'chiring\n"
                "4️⃣ Faylni botga yuboring → avtomatik yuklanadi"
            )
        )
        return

    if data == "adm_import":
        context.user_data["adm_step"] = "excel_import"
        await query.message.edit_text(
            "📥 Excel faylni yuboring (kiym_shablon.xlsx):",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Bekor", callback_data="adm_products")]
            ])
        )
        return

    if data == "adm_add":
        context.user_data["adm_step"] = "add_gender"
        context.user_data["new_product"] = {}
        await query.message.edit_text(
            "➕ Yangi mahsulot\n\nJins tanlang:",
            reply_markup=admin_add_gender_kb()
        )
        return

    if data.startswith("adm_gender_"):
        gender = "O'g'il" if "ogil" in data else "Qiz"
        context.user_data["new_product"]["gender"] = gender
        context.user_data["adm_step"] = "add_origin"
        await query.message.edit_text(
            f"➕ Yangi mahsulot\n✅ Jins: {gender}\n\nFabrika tanlang:",
            reply_markup=admin_add_origin_kb()
        )
        return

    if data.startswith("adm_origin_"):
        origin = data[len("adm_origin_"):]
        context.user_data["new_product"]["origin"] = origin
        context.user_data["adm_step"] = "add_season"
        context.user_data["new_product"]["seasons"] = []
        p = context.user_data["new_product"]
        await query.message.edit_text(
            f"➕ Yangi mahsulot\n✅ Jins: {p['gender']}\n✅ Fabrika: {origin}\n\nFasl tanlang (bir nechta bo'lishi mumkin):",
            reply_markup=admin_add_season_kb([])
        )
        return

    if data.startswith("adm_season_") and data != "adm_season_done":
        season = data[len("adm_season_"):]
        seasons = context.user_data["new_product"].get("seasons", [])
        if season in seasons:
            seasons.remove(season)
        else:
            seasons.append(season)
        context.user_data["new_product"]["seasons"] = seasons
        p = context.user_data["new_product"]
        await query.message.edit_reply_markup(
            reply_markup=admin_add_season_kb(seasons)
        )
        return

    if data == "adm_season_done":
        seasons = context.user_data["new_product"].get("seasons", [])
        if not seasons:
            await query.answer("Kamida 1 ta fasl tanlang!", show_alert=True)
            return
        context.user_data["adm_step"] = "add_category"
        p = context.user_data["new_product"]
        await query.message.edit_text(
            f"➕ Yangi mahsulot\n✅ Jins: {p['gender']}\n✅ Fabrika: {p['origin']}\n✅ Fasl: {', '.join(seasons)}\n\nKategoriya tanlang:",
            reply_markup=admin_add_category_kb()
        )
        return

    if data.startswith("adm_cat_"):
        category = data[len("adm_cat_"):]
        context.user_data["new_product"]["category"] = category
        context.user_data["adm_step"] = "add_photo_num"
        p = context.user_data["new_product"]
        await query.message.edit_text(
            f"➕ Yangi mahsulot\n"
            f"✅ Jins: {p['gender']}\n"
            f"✅ Fabrika: {p['origin']}\n"
            f"✅ Fasl: {', '.join(p['seasons'])}\n"
            f"✅ Kategoriya: {category}\n\n"
            f"📸 Rasm raqamini yozing (masalan: 3)\n"
            f"Rasm yo'q bo'lsa — avval 📸 Rasm yuklash bo'limiga boring",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Orqaga", callback_data="adm_add")]
            ])
        )
        return

    if data == "adm_list_0" or data.startswith("adm_list_"):
        page = int(data.split("_")[2]) if "_" in data[8:] else 0
        offset = page * 5
        cur.execute("SELECT id, name, size, price, count FROM shop_products ORDER BY id DESC LIMIT 5 OFFSET %s", (offset,))
        rows = cur.fetchall()
        cur.execute("SELECT COUNT(*) FROM shop_products")
        total = cur.fetchone()[0]
        if not rows:
            await query.message.edit_text(
                "📦 Mahsulotlar yo'q",
                reply_markup=admin_products_kb()
            )
            return
        text = f"📦 Mahsulotlar ({total} ta):\n\n"
        kb = []
        for r in rows:
            text += f"🆔{r[0]} | {r[1]} | {r[2]}sm | {r[3]} | soni:{r[4]}\n"
            kb.append([
                InlineKeyboardButton(f"#{r[0]} {r[1]}", callback_data=f"adm_product_{r[0]}_{page}"),
            ])
        nav = []
        if page > 0:
            nav.append(InlineKeyboardButton("⬅️", callback_data=f"adm_list_{page-1}"))
        if offset + 5 < total:
            nav.append(InlineKeyboardButton("➡️", callback_data=f"adm_list_{page+1}"))
        if nav:
            kb.append(nav)
        kb.append([InlineKeyboardButton("🔙 Orqaga", callback_data="adm_products")])
        await query.message.edit_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith("adm_product_"):
        parts = data.split("_")
        product_id = int(parts[2])
        page = int(parts[3]) if len(parts) > 3 else 0
        p = next((x for x in products if x["id"] == product_id), None)
        if not p:
            await query.answer("Topilmadi", show_alert=True)
            return
        cur.execute("SELECT COUNT(*) FROM shop_products")
        total = cur.fetchone()[0]
        has_prev = page > 0
        has_next = (page + 1) * 5 < total
        text = (
            f"📦 Mahsulot #{product_id}\n\n"
            f"📛 {p['name']}\n"
            f"👤 {p['gender']}\n"
            f"🏭 {p['origin']}\n"
            f"🌸 {p['season']}\n"
            f"📂 {p['category']}\n"
            f"📏 {p['size']} sm\n"
            f"💰 {p['price']}\n"
            f"📦 Soni: {p['count']}\n"
            f"🔒 Band: {p['reserved']}"
        )
        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✏️ Tahrir", callback_data=f"edit_{product_id}"),
                InlineKeyboardButton("🗑 O'chir", callback_data=f"delete_{product_id}"),
            ],
            [InlineKeyboardButton("🔙 Ro'yxat", callback_data=f"adm_list_{page}")],
        ])
        if p.get("photo"):
            try:
                await query.message.delete()
                await query.message.chat.send_photo(
                    photo=p["photo"],
                    caption=text,
                    reply_markup=kb
                )
            except Exception:
                await query.message.edit_text(text, reply_markup=kb)
        else:
            await query.message.edit_text(text, reply_markup=kb)
        return

    if data == "adm_clear":
        await query.message.edit_text(
            "⚠️ Barcha mahsulotlarni o'chirishni tasdiqlaysizmi?",
            reply_markup=admin_clear_confirm_kb()
        )
        return

    if data == "adm_save":
        p = context.user_data.get("new_product", {})
        if not p.get("photo") or not p.get("name"):
            await query.answer("Ma'lumotlar to'liq emas!", show_alert=True)
            return
        try:
            cur.execute("""
                INSERT INTO shop_products
                    (photo, gender, origin, season, category, name, size, price, count, reserved, cost)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s)
            """, (
                p["photo"],
                p.get("gender", ""),
                p.get("origin", ""),
                ",".join(p.get("seasons", [])),
                p.get("category", ""),
                p.get("name", ""),
                p.get("size", ""),
                p.get("price", ""),
                p.get("count", 0),
                p.get("cost", 0),
            ))
            conn.commit()
            load_products_from_db()
            context.user_data.pop("new_product", None)
            context.user_data.pop("adm_step", None)
            await query.message.edit_text(
                f"✅ Mahsulot qo'shildi!\n📛 {p['name']}",
                reply_markup=InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("➕ Yana qo'shish", callback_data="adm_add"),
                        InlineKeyboardButton("📋 Ro'yxat",       callback_data="adm_list_0"),
                    ],
                    [InlineKeyboardButton("🔙 Admin panel", callback_data="adm_back_main")],
                ])
            )
        except Exception as e:
            await query.answer(f"Xato: {e}", show_alert=True)
        return

# ===== FILTER BLOK =====
    if data.startswith("g_") or data.startswith("o_") or data.startswith("s_"):

        if data.startswith("g_"):
            g = data[2:]

            if "o‘g" in g or "og" in g:
                context.user_data["filter_gender"] = "o‘g‘il"
            else:
                context.user_data["filter_gender"] = "qiz"

        elif data.startswith("o_"):
            context.user_data["filter_origin"] = data[2:]

        elif data.startswith("s_"):
            context.user_data["filter_season"] = data[2:]

        # 🔥 TEXTNI YANGILAYMIZ
        await query.message.edit_text(
            f"🔎 Tanlang:\n\n"
            f"Jins: {context.user_data.get('filter_gender','-')}\n"
            f"Fabrika: {context.user_data.get('filter_origin','-')}\n"
            f"Fasl: {context.user_data.get('filter_season','-')}\n",
            reply_markup=get_filter_menu(context.user_data)
        )
        return
    # ===== RESET =====
    if data == "reset":
        context.user_data.clear()

        await query.message.edit_text(
            "🔎 Tanlang:\n\nJins: -\nFabrika: -\nFasl: -\nRazmer: -",
            reply_markup=get_filter_menu(context.user_data)
        )
        return


    # ===== APPLY =====
    if data == "apply":

        context.user_data["step"] = "inline_category"

        keyboard = get_category_buttons(context)

        await query.message.reply_text(
            "📂 Kategoriya tanlang:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return
    elif data.startswith("add_"):
        
        product_id = int(data.split("_")[1])

        if product_id not in product_locks:
            product_locks[product_id] = asyncio.Lock()

        async with product_locks[product_id]:

            product = next((x for x in products if x["id"] == product_id), None)
            if not product:
                await query.answer("❌ Topilmadi", show_alert=True)
                return

            # 🔥 REAL TEKSHIRUV
            available = product["count"] - product.get("reserved", 0)

            if available <= 0:
                await query.answer("❌ Bu mahsulot band yoki qolmagan", show_alert=True)
                return

            if user_id not in carts:
                carts[user_id] = {}

            if product_id in carts[user_id]:
                carts[user_id][product_id]["qty"] += 1
            else:
                carts[user_id][product_id] = {
                    "qty": 1,
                    "time": time.time()
                }

            carts[user_id][product_id]["time"] = time.time()

            # 🔥 ENG MUHIM — RESERVE
            product["reserved"] = product.get("reserved", 0) + 1

        await query.answer("✅ Savatga qo‘shildi")

        keyboard = [
            [InlineKeyboardButton("🧺 Savatga o‘tish", callback_data="go_cart")]
        ]

        await query.message.reply_text(
            "🛒 Mahsulot vaqtincha siz uchun band qilindi!\n⏳ 2 soat ichida xarid qilmasangiz o‘chiriladi.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data.startswith("edit_"):
        product_id = int(data.split("_")[1])

        old_product = next((x for x in products if x["id"] == product_id), None)
        if not old_product:
            await query.message.reply_text("❌ Topilmadi")
            return

        # eski mahsulotni o‘chir
        cur.execute("DELETE FROM shop_products WHERE id=%s", (product_id,))
        conn.commit()
        load_products_from_db()

        # bot rasmni chiqaradi
        await context.bot.send_photo(
            chat_id=query.message.chat_id,
            photo=old_product["photo"]
        )

        # 🔥 ENG MUHIM — bot yuborgan rasmni ham ishlatamiz
        await start_photo_flow(context, query.message.chat_id, old_product["photo"])
    elif data.startswith("delete_"):
        if query.from_user.id != ADMIN_ID:
            return

        product_id = int(data.split("_")[1])

        cur.execute("DELETE FROM shop_products WHERE id=%s", (product_id,))
        conn.commit()

        load_products_from_db()

        await query.message.reply_text("✅ Mahsulot o‘chirildi")
    elif data == "clear_yes":
        if query.from_user.id != ADMIN_ID:
            return

        cur.execute("DELETE FROM shop_products")
        conn.commit()
        load_products_from_db()
       # save_products()

        await query.message.reply_text("✅ Barcha mahsulotlar o‘chirildi")

    elif data == "clear_no":
        await query.message.reply_text("❌ Bekor qilindi")

            # 🔥 oldinga orqaga
    elif data == "next_one" or data == "prev_one":

        # 🔄 indexni o‘zgartiramiz
        if data == "next_one":
            context.user_data["i"] += 1
            if context.user_data["i"] >= len(context.user_data["filtered"]):
                context.user_data["i"] = len(context.user_data["filtered"]) - 1

        else:  # prev_one
            context.user_data["i"] -= 1
            if context.user_data["i"] < 0:
                context.user_data["i"] = 0

        # 🔥 MAHSULOTNI OLAMIZ
        p = context.user_data["filtered"][context.user_data["i"]]

        # 🔘 TUGMALAR
        keyboard = []
        nav = []

        if context.user_data["i"] > 0:
            nav.append(InlineKeyboardButton("⬅️", callback_data="prev_one"))

        if context.user_data["i"] < len(context.user_data["filtered"]) - 1:
            nav.append(InlineKeyboardButton("➡️", callback_data="next_one"))

        if nav:
            keyboard.append(nav)

        keyboard.append([
            InlineKeyboardButton("🛒 Savatga qo‘shish", callback_data=f"add_{p.get('id')}")
        ])

        if update.effective_user.id == ADMIN_ID:
            keyboard.append([
                InlineKeyboardButton("✏️ Edit", callback_data=f"edit_{p.get('id')}"),
                InlineKeyboardButton("🗑 O‘chirish", callback_data=f"delete_{p.get('id')}")
            ])

        from telegram import InputMediaPhoto

        await query.message.edit_media(
            media=InputMediaPhoto(
                media=p.get("photo"),
                caption=f"{context.user_data['i']+1}/{len(context.user_data['filtered'])}\n\n"
                        f"{p.get('name')}\n📏 {p.get('size')}\n💰 {p.get('price')}"
            ),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data.startswith("plus_"):
        product_id = int(data.split("_")[1])

        product = next((x for x in products if x["id"] == product_id), None)
        if not product:
            return

        if product["count"] - product.get("reserved", 0) <= 0:
            await query.answer("❌ Yetarli mahsulot yo‘q", show_alert=True)
            return

        if user_id not in carts:
            carts[user_id] = {}

        # 🔥 ENG MUHIM TUZATISH
        if product_id in carts[user_id]:
            carts[user_id][product_id]["qty"] += 1
        else:
            carts[user_id][product_id] = {
                "qty": 1,
                "time": time.time()
            }

        carts[user_id][product_id]["time"] = time.time()
        product["reserved"] += 1

        await query.answer("➕ Qo‘shildi")
    elif data.startswith("send_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USERGA
        await context.bot.send_message(
            chat_id=user_id,
            text="🚚 Buyurtmangiz yo‘lga chiqdi!\n⏳ 1 soat ichida yetkaziladi."
        )

        # ADMINGA
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"🚚 Buyurtma jo‘natildi\nID: {order_id}"
        )

        await query.answer("Yuborildi")

 
    elif data.startswith("deliver_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USERGA
        await context.bot.send_message(
            chat_id=user_id,
            text="🚚 Buyurtmangiz yetkazilmoqda!\n⏳ 1 soat ichida yetkaziladi."
        )

        # ADMINGA
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"🚚 Yetkazish boshlandi\nID: {order_id}"
        )

        await query.answer("Yuborildi")

    elif data.startswith("done_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="✅ Buyurtmangizni qabul qilib oldingiz. Rahmat 😊"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"✅ YAKUNLANDI\nID: {order_id}"
        )

        cur.execute("DELETE FROM shop_orders WHERE id=%s", (order_id,))
        conn.commit()
       # save_orders()

        await query.answer("Yakunlandi")

    elif data.startswith("ready_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📦 Buyurtmangiz tayyor!"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 TAYYOR\nID: {order_id}"
        )

        await query.answer("Tayyor qilindi")

    elif data.startswith("cancel_"):
        order_id = data.split("_")[1]
        cur.execute("SELECT user_id, cart FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]
        cart = json.loads(row[1])

        # 🔥 mahsulotni qaytaramiz
        for product_id, item in cart.items():
            qty = item["qty"]

            p = next((x for x in products if x["id"] == int(product_id)), None)

            if p:
                p["count"] += qty
                p["reserved"] = max(0, p["reserved"] - qty)

        load_products_from_db()  # 🔥 ENG MUHIM

        carts[user_id] = {}

        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Buyurtmangiz bekor qilindi"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"❌ BUYURTMA BEKOR QILINDI\nID: {order_id}"
        )

        cur.execute("DELETE FROM shop_orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("Bekor qilindi")
    elif data.startswith("user_cancel_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id, cart FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            await query.answer("Allaqachon bekor qilingan")
            return

        user_id = row[0]
        cart = json.loads(row[1])

        # 🔥 mahsulotni qaytaramiz
        for pid, item in cart.items():
            qty = item["qty"]
            p = next((x for x in products if x["id"] == int(pid)), None)
            if p:
                p["count"] += qty
                p["reserved"] = max(0, p.get("reserved", 0) - qty)

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"❌ Mijoz bekor qildi\nID: {order_id}"
        )

        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Buyurtma bekor qilindi"
        )

        cur.execute("DELETE FROM shop_orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("Bekor qilindi")

    elif data.startswith("delivered_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USER ga
        await context.bot.send_message(
            chat_id=user_id,
            text=f"📦 Buyurtma yetkazildi!\n🆔 ID: {order_id}"
        )

        # ADMIN ga
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 Yetkazildi\nID: {order_id}"
        )

        await query.answer("Yetkazildi")
    elif data.startswith("paid_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # USER ga
        await context.bot.send_message(
            chat_id=user_id,
            text="💰 To‘lov qabul qilindi! Rahmat 😊"
        )

        # ADMIN ga
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"💰 To‘lov olindi\nID: {order_id}"
        )

        # ORDERNI O‘CHIRAMIZ
        cur.execute("DELETE FROM shop_orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("To‘lov olindi")

    elif data.startswith("accept_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        # 🔥 ADMIN MA’LUMOTI
        ADMIN_PHONE = "+998915388499"
        ADDRESS = "Samarqand vil. Pastdarg'om tum. charxin shax. charos ko‘chasi 53 uy Adminlar aloqaga chiqishadi va olib ketish vaqtini keishiladi."

        #LAT = 39.690149
        #LON = 66.824828

        # USERGA
        await context.bot.send_message(
            chat_id=user_id,
            text=f"✅ Buyurtmangiz qabul qilindi!\n\n📞 Tel: {ADMIN_PHONE}\n🏠 Manzil: {ADDRESS}"
        )

      #  await context.bot.send_location(
       #     chat_id=user_id,
        #    latitude=LAT,
         #   longitude=LON
        #)

        # ADMIN ga
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"✅ BUYURTMA QABUL QILINDI\nID: {order_id}"
        )

        await query.answer("Yuborildi")
    elif data.startswith("contact_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📞 Admin sizga 10 min ichida bog‘lanadi"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📞 Mijoz bilan bog‘laning\nID: {order_id}"
        )

        await query.answer("Yuborildi")
    elif data.startswith("picked_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📦 Buyurtma yakunlandi! Rahmat 😊"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 BUYURTMA YAKUNLANDI\nID: {order_id}"
        )

        # 🔥 ORDERNI O‘CHIRAMIZ
        cur.execute("DELETE FROM shop_orders WHERE id=%s", (order_id,))
        conn.commit()

        await query.answer("Yakunlandi")

    elif data == "go_cart":
        user_id = query.from_user.id
        cart = carts.get(user_id, {})

        
        now = time.time()
        new_cart = {}

        # ⏳ 2 soatdan eski itemlarni tozalash
        for pid, item in cart.items():
            if now - item["time"] < 7200:
                new_cart[pid] = item
            else:
                qty = item["qty"]
                p = next((x for x in products if x["id"] == int(pid)), None)
                if p:
                    p["reserved"] = max(0, p.get("reserved", 0) - qty)

        carts[user_id] = new_cart
        cart = new_cart

        # 🧺 Savat bo‘sh bo‘lsa
        if not cart:
            await query.message.reply_text("🧺 Savat bo‘sh")
            return

        msg = "🧺 Savat:\n\n"
        total = 0
        keyboard = []

        # 📦 Savat ichidagi mahsulotlar
        for pid, item in cart.items():
            try:
                pid = int(pid)
            except:
                continue

            p = next((x for x in products if x["id"] == pid), None)
            if not p:
                continue

            qty = item["qty"]

            # 💰 narxni tozalab olish
            price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

            summa = price * qty
            total += summa

            msg += f"{p['name']} x{qty} = {summa}\n"

            keyboard.append([
                InlineKeyboardButton(f"❌ {p['name']}", callback_data=f"del_{pid}")
            ])

        msg += f"\n💰 Jami: {total}"

        keyboard.append([
            InlineKeyboardButton("🚚 Buyurtma berish", callback_data="checkout")
        ])
        keyboard.append([
            InlineKeyboardButton("🔙 Orqaga", callback_data="back")
        ])

        await query.message.reply_text(
            msg,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    elif data.startswith("del_"):
        product_id = int(data.split("_")[1])

        cart = carts.get(user_id, {})

        # 🔥 cart kaliti int yoki str bo'lishi mumkin — ikkalasini tekshiramiz
        cart_key = product_id if product_id in cart else str(product_id) if str(product_id) in cart else None

        if cart_key is not None:
            qty = cart[cart_key]["qty"]

            p = next((x for x in products if x["id"] == product_id), None)
            if p:
                p["reserved"] = max(0, p.get("reserved", 0) - qty)

            cart.pop(cart_key)

        await query.answer("❌ O‘chirildi")

        if not cart:
            await query.message.reply_text("🧺 Savat bo‘sh")
            return

        msg = "🧺 Savat:\n\n"
        total = 0
        keyboard = []

        for pid, item in cart.items():
            qty = item["qty"]

            p = next((x for x in products if x["id"] == int(pid)), None)
            if not p:
                continue

            price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

            summa = price * qty
            total += summa

            msg += f"{p['name']} x{qty} = {summa}\n"

            keyboard.append([
                InlineKeyboardButton("❌", callback_data=f"del_{pid}")
            ])

        msg += f"\n💰 Jami: {total}"

        keyboard.append([InlineKeyboardButton("🚚 Buyurtma berish", callback_data="checkout")])
        keyboard.append([InlineKeyboardButton("🔙 Orqaga", callback_data="back")])

        await query.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard))
    elif data == "checkout":

        context.user_data["order_started"] = time.time()
        context.user_data["order_step"] = "choose_type"

        keyboard = [
            ["🚚 Dastavka", "📍 Olib ketish"],
            ["🏠 Bosh menyu"]
        ]

        await query.message.reply_text(
            "Qanday olasiz?",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )


    elif data == "back":
        await query.message.reply_text(
            "🏠 Bosh menyu",
            reply_markup=MAIN_MENU
            
        )
    elif data.startswith("confirm_"):
        order_id = data.split("_")[1]

        cur.execute("SELECT user_id FROM shop_orders WHERE id=%s", (order_id,))
        row = cur.fetchone()

        if not row:
            return

        user_id = row[0]

        await context.bot.send_message(
            chat_id=user_id,
            text="📦 Buyurtmangiz tayyor! Kelishilgan vaqtda olib ketishingiz mumkin"
        )

        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"📦 BUYURTMA TASDIQLANDI\nID: {order_id}"
        )

        await query.answer("Tasdiqlandi")
async def location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if context.user_data.get("order_step") != "location":
        return

    location = update.message.location
    lat = location.latitude
    lon = location.longitude

    user_id = update.effective_user.id
    cart = carts.get(user_id, {})

    if not cart:
        await update.message.reply_text("❌ Savat bo‘sh")
        return

    total = 0

    for pid, item in cart.items():
        qty = item["qty"]
        p = next((x for x in products if x["id"] == int(pid)), None)
        if not p:
            continue

        price = int(''.join(filter(str.isdigit, str(p.get("price", 0)))))

        total += price * qty   # ✅ TO‘G‘RI

    delivery = 0
    final = total + delivery

    context.user_data["temp_order"] = {
        "cart": cart,
        "location": {"lat": lat, "lon": lon},
        "total": final,
        "type": context.user_data.get("order_type")
    }

    await update.message.reply_text("⏳ Lokatsiya qabul qilindi, hisoblanmoqda...")

    context.user_data["order_step"] = "phone"

    keyboard = [
        [KeyboardButton("📞 Telefon yuborish", request_contact=True)],
        ["🏠 Bosh menyu"]
    ]

    await update.message.reply_text(
        "📞 Telefon raqamingizni yuboring yoki yozing (+998...):",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )
async def contact_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("order_step") != "phone":
        return

    contact = update.message.contact
    phone = contact.phone_number

    if not phone.startswith("+"):
        phone = "+" + phone

    data = context.user_data.get("temp_order")
    if not data:
        await update.message.reply_text("❌ Xatolik")
        return

    user_id = update.effective_user.id
    cur.execute("""
    INSERT INTO shop_orders (user_id, cart, location, phone, total, status, time)
    VALUES (%s,%s,%s,%s,%s,%s,%s)
    RETURNING id
    """, (
        user_id,
        json.dumps(data["cart"]),
        json.dumps(data["location"]),
        phone,
        data["total"],
        "new",
        time.time()
    ))

    order_id = str(cur.fetchone()[0])
    conn.commit()

    # ===== MAHSULOTNI KAMAYTIRISH =====
    for pid, item in data["cart"].items():
        qty = item["qty"]
        p = next((x for x in products if x["id"] == int(pid)), None)
        if p:
            p["count"] -= qty
            p["reserved"] = max(0, p.get("reserved", 0) - qty)
            # 🔥 DB ga ham yozamiz
            cur.execute(
                "UPDATE shop_products SET count = count - %s, reserved = GREATEST(0, reserved - %s) WHERE id = %s",
                (qty, qty, p["id"])
            )
    conn.commit()

    # ===== USERGA MAHSULOT =====
    for pid, item in data["cart"].items():
        p = next((x for x in products if x["id"] == int(pid)), None)
        if not p:
            continue
        qty = item["qty"]

        await context.bot.send_photo(
            chat_id=user_id,
            photo=p["photo"],
            caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
        )

    # ===== ADMINGA MAHSULOT =====
    for pid, item in data["cart"].items():
        p = next((x for x in products if x["id"] == int(pid)), None)
        if not p:
            continue
        qty = item["qty"]

        await context.bot.send_photo(
            chat_id=ADMIN_ID,
            photo=p["photo"],
            caption=f"{p['name']}\n📏 Razmer: {p['size']}\n💰 {p['price']} x{qty}"
        )

    # ===== ADMIN TUGMALAR =====
    if data.get("type") == "delivery":
        admin_keyboard = [
            [InlineKeyboardButton("📞 Aloqa", callback_data=f"contact_{order_id}")],
            [InlineKeyboardButton("🚚 Buyurtmani jo'natish", callback_data=f"send_{order_id}")],
            [InlineKeyboardButton("✅ Yakunlandi", callback_data=f"done_{order_id}")],
            [InlineKeyboardButton("❌ Bekor", callback_data=f"cancel_{order_id}")]
        ]
    else:
        admin_keyboard = [
            [InlineKeyboardButton("📞 Aloqa", callback_data=f"contact_{order_id}")],
            [InlineKeyboardButton("📦 Buyurtmani tasdiqlash", callback_data=f"confirm_{order_id}")],
            [InlineKeyboardButton("✅ Yakunlandi", callback_data=f"done_{order_id}")],
            [InlineKeyboardButton("❌ Bekor", callback_data=f"cancel_{order_id}")]
        ]

    # ===== TEXT =====
    if data.get("type") == "delivery":
        if data.get("location") and "lat" in data["location"]:
            lat = data["location"]["lat"]
            lon = data["location"]["lon"]
            loc = f"\n📍 https://maps.google.com/?q={lat},{lon}"
        else:
            loc = "\n📍 Lokatsiya yuborilmadi"

        text_admin = (
            f"🚚 DASTAVKA\n"
            f"📞 {phone}\n"
            f"💰 {data['total']}{loc}"
        )

        await update.message.reply_text(
            "🚚 Buyurtma qabul qilindi!\n📞 Admin siz bilan tez orada bog'lanadi.",
            reply_markup=MAIN_MENU
        )
    else:
        text_admin = (
            f"📍 OLIB KETISH\n"
            f"📞 {phone}\n"
            f"💰 {data['total']}\n"
            f"🏠 Samarqand, Pastdarg'om, Charxin\n"
        )

        await update.message.reply_text(
            "📍 Olib ketish manzili:\nSamarqand, Pastdarg'om, Charxin\n A'loqa 📞 +998915388499  Adminlar o'zlari a'loqaga chiqishadi va manzilni yetgazishadi. "
        )

        await update.message.reply_text(
            "🏠 Bosh menyu",
            reply_markup=MAIN_MENU
        )

    # 🔥 ENG MUHIM — ADMIN GA HAR DOIM YUBORILADI
    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=text_admin,
        reply_markup=InlineKeyboardMarkup(admin_keyboard)
    )

    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=f"📦 Pickup tayyor\nID: {order_id}"
    )

    # ===== TOZALASH =====
    carts[user_id] = {}
    context.user_data.clear()

# Application'ni qurishda quyidagi tartibda qo'shing:



# ─────────────────────────────────────────────
# SHABLON GENERATSIYA (openpyxl)
# ─────────────────────────────────────────────
def generate_shablon_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = [
        ("Rasm\nraqami", 10, True),
        ("Nomi",           22, True),
        ("Jins",           12, True),
        ("Fabrika",        16, True),
        ("Fasl",           14, True),
        ("Kategoriya",     22, True),
        ("Razmer\n(sm)",  10, True),
        ("Narx\n(so'm)", 12, True),
        ("Tannarx",        12, False),
        ("Soni",           8,  True),
        ("Ranglar",        18, False),
        ("Izoh",           20, False),
    ]

    QIYMATLAR = [
        "Raqam (1, 2, 3...)",
        "Istalgan nom",
        "O'g'il  |  Qiz",
        "Vodiy  |  Xitoy  |  Turkiya  |  8-mart fabrika",
        "Yozgi  |  Qishki  |  Bahor  |  Kuz",
        "2 talik kiyim  |  3 talik kiyim  |  futbolka  |  shim  |  qalin kiyim  |  shortik  |  oyoq kiyim  |  bosh kiyim  |  ichki kiyim",
        "44  yoki  86-92",
        "Raqam: 85000",
        "Raqam: 60000",
        "Raqam: 4",
        "Oq:2,Qora:3",
        "Istalgan matn",
    ]

    NAMUNA = ["1", "Bolalar kostyumi", "O'g'il", "Xitoy",
              "Yozgi", "2 talik kiyim", "86-92", "85000", "60000", "4", "Oq:2", ""]

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 20

    for ci, (title, width, req) in enumerate(COLS, 1):
        L = get_column_letter(ci)
        ws.column_dimensions[L].width = width

        c = ws.cell(row=1, column=ci, value=title)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

        n = ws.cell(row=2, column=ci, value=QIYMATLAR[ci-1])
        n.font = Font(italic=True, color="1F4E79", size=8, bold=True)
        n.fill = PatternFill("solid", start_color="DEEAF1")
        n.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n.border = border

        ex = ws.cell(row=3, column=ci, value=NAMUNA[ci-1])
        ex.font = Font(italic=True, color="7B6000", size=10)
        ex.fill = PatternFill("solid", start_color="FFF2CC")
        ex.border = border
        ex.alignment = Alignment(vertical="center")

        for row in range(4, 104):
            cell = ws.cell(row=row, column=ci)
            cell.fill = PatternFill("solid", start_color="D9E1F2" if req else "EBF5FB")
            cell.border = border
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center")

    ws.cell(row=3, column=1).value = "NAMUNA — o'chiring"
    ws.cell(row=3, column=1).font = Font(bold=True, color="7B6000", size=9)

    ws.add_data_validation(DataValidation(
        type="list", formula1='"O\'g\'il,Qiz"',
        allow_blank=True, showDropDown=False, sqref="C4:C103"))

    ws.add_data_validation(DataValidation(
        type="list", formula1='"Vodiy,Xitoy,Turkiya,8-mart fabrika"',
        allow_blank=True, showDropDown=False, sqref="D4:D103"))

    ws.add_data_validation(DataValidation(
        type="list", formula1='"Yozgi,Qishki,Bahor,Kuz"',
        allow_blank=True, showDropDown=False, sqref="E4:E103"))

    ws.add_data_validation(DataValidation(
        type="list",
        formula1='"2 talik kiyim,3 talik kiyim,futbolka,shim,qalin kiyim,shortik,oyoq kiyim,bosh kiyim,ichki kiyim"',
        allow_blank=True, showDropDown=False, sqref="F4:F103"))

    ws.add_data_validation(DataValidation(
        type="whole", operator="greaterThan", formula1="0",
        allow_blank=True, sqref="H4:J103"))

    ws.freeze_panes = "A4"

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def shablon_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    await update.message.reply_text("⏳ Shablon tayyorlanmoqda...")
    data = generate_shablon_bytes()
    await update.message.reply_document(
        document=data,
        filename="kiym_shablon.xlsx",
        caption=(
            "📋 Excel shablon\n\n"
            "1️⃣ /get_id → rasmlarni yuboring → file_id larni oling\n"
            "2️⃣ Shu faylni to'ldiring (dropdown lar bor)\n"
            "3️⃣ NAMUNA qatorini o'chiring\n"
            "4️⃣ Faylni menga yuboring → DB ga yozaman"
        )
    )

app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("get_id", get_id_command))
app.add_handler(CommandHandler("done", done_command))
app.add_handler(CommandHandler("shablon", shablon_command))
app.add_handler(MessageHandler(filters.Document.ALL, excel_import_handler))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))
app.add_handler(MessageHandler(filters.PHOTO, photo_handler))
app.add_handler(MessageHandler(filters.CONTACT, contact_handler))
app.add_handler(MessageHandler(filters.LOCATION, location_handler))
app.add_handler(CallbackQueryHandler(button_handler, pattern=".*"))
load_products_from_db()
app.run_polling()
